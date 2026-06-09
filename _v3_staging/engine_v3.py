"""
═══════════════════════════════════════════════════════════════════════════════
  SMA OUTFIT DETECTION ENGINE V3
  Element 47 — Level Significance + Sequential Conditional
═══════════════════════════════════════════════════════════════════════════════

  Engine 3 builds on engines 1 and 2 with three core advances:

  1. LEVEL SIGNIFICANCE SCORING
     Replaces raw hit count with a 7-factor composite:
     hit_rate × volume_ratio × hold_rate × cross_outfit × cross_tf
     × recency × persistence

  2. SEQUENTIAL CONDITIONAL (PARM Logic)
     For every signal, checks the current candle vs the PARM (key variable SMA):
     - Close BELOW PARM + micro-term same-outfit active = HOLD
     - Close BELOW PARM + no micro-term                = IGNORE
     - Close AT/ABOVE PARM + micro-term active         = STRONG
     - Close AT/ABOVE PARM + no micro-term             = WEAK

  3. TIMEFRAME-EQUAL SCORING
     Inherited from engine 2 — deciseconds normalized by TF baseline.

  Uses the same Webull data adapter, outfits, and hit detection as engines 1/2.
  Completely separate codebase — does not modify either prior engine.

  Commands:
    docker exec e47_engine_v3 python3 /app/engine_v3.py --source webull --xlsx
    docker exec e47_engine_v3 python3 /app/engine_v3.py --source webull --top-n 500

  Output: /cache/output/v3/
═══════════════════════════════════════════════════════════════════════════════
"""

from __future__ import annotations

import os
import sys
import time
import uuid
import json
import hmac
import base64
import hashlib
import logging
import argparse
import asyncio
import math
import multiprocessing
import random
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta, timezone
from itertools import product
from typing import Optional, Iterable

import numpy as np
import pandas as pd
import requests

import sys as _sys_early
_sys_early.path.insert(0, "/app_v3")
_sys_early.path.insert(0, "/app")

from scoring_v3 import (
    LevelScore, score_entry,
    compute_hit_rate, compute_volume_ratio, compute_hold_rate,
    compute_recency, compute_cross_outfit_score, compute_cross_tf_score,
    compute_persistence_score,
)
from conditions_v3 import (
    ConditionalState, evaluate_condition,
    filter_actionable, summarize_states,
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ── Import outfits, systems, universes, clients from engine_normalized ────────
# Engine v3 reuses all the same data — outfits, tickers, Webull adapter.
# We import from the sma_engine directory which is mounted at /app.
sys.path.insert(0, "/app")
sys.path.insert(0, "/app_v3")
try:
    from engine_normalized import (
        OUTFITS, SYSTEMS, TIMEFRAMES_STANDARD,
        WEBULL_TIMESPAN_MAP, TF_MINUTES, TF_DECISECONDS,
        UNIVERSE_TIER_1, UNIVERSE_TIER_2, UNIVERSE_TIER_3,
        WebullClient, MockClient, DataClient,
        compute_smas, Hit, HashMapEntry, HashMapStore,
        detect_hits, freshness_score,
        _scan_worker_fn, _scan_worker_init,
        SystemState, evaluate_systems,
        best_offset,
    )
    logging.info("Loaded base components from engine_normalized")
except ImportError as e:
    logging.error(f"Could not import from engine_normalized: {e}")
    logging.error("Make sure /app contains engine_normalized.py")
    sys.exit(1)


# ── Output directory ──────────────────────────────────────────────────────────
OUTPUT_DIR = "/cache/output/v3"

# ── Micro-term timeframes for conditional check ───────────────────────────────
MICRO_TFS = {"1m", "5m"}


# ═══════════════════════════════════════════════════════════════════════════════
# V3 RANKING — level significance + normalized deciseconds
# ═══════════════════════════════════════════════════════════════════════════════

def rank_entries_v3(
    store: HashMapStore,
    candle_cache: dict,
    now: pd.Timestamp,
    all_entries_flat: list[dict],
    cumulative_ds: Optional[dict],
    lookback: int,
    weight_freshness: float = 0.3,
    min_tf_minutes: int = 0,
) -> list[tuple[HashMapEntry, LevelScore]]:
    """
    Rank all entries using the v3 composite level significance score.
    Returns list of (entry, LevelScore) sorted by composite descending.
    """
    scored = []

    for e in store.all():
        tf_mins = TF_MINUTES.get(e.timeframe, 5)
        if min_tf_minutes > 0 and tf_mins < min_tf_minutes:
            continue

        df = candle_cache.get((e.ticker, e.timeframe), pd.DataFrame())

        try:
            ls = score_entry(
                entry          = e,
                df             = df,
                all_entries    = all_entries_flat,
                cumulative_ds  = cumulative_ds,
                lookback       = lookback,
            )
        except Exception as ex:
            logging.debug(f"Scoring failed for {e.ticker} {e.timeframe}: {ex}")
            continue

        # Apply freshness on top of composite (same as original engine)
        f = freshness_score(e.last_hit_ts, now, tf_mins)
        ls.composite = round(ls.composite * (1 + weight_freshness * f), 2)
        ls.outfit_name = next(
            (o["name"] for o in OUTFITS if o["id"] == e.outfit_id), ""
        )

        scored.append((e, ls))

    return sorted(scored, key=lambda x: x[1].composite, reverse=True)


# ═══════════════════════════════════════════════════════════════════════════════
# V3 ENGINE CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class EngineV3Config:
    universe:           list[str]
    timeframes:         list[str]   = field(default_factory=lambda: [
        "1m", "5m", "15m", "30m", "1h", "2h", "4h", "1d", "1w", "1mo"
    ])
    outfits:            list[dict]  = field(default_factory=lambda: OUTFITS)
    lookback:           int         = 390
    candle_count:       int         = 999
    hit_mode:           str         = "exact"
    hit_tolerance:      float       = 0.0
    refresh_bars:       int         = 20
    min_tf_minutes:     int         = 0
    weight_freshness:   float       = 0.3

    # V3-specific
    run_conditions:     bool        = True   # run sequential conditional layer
    min_composite:      float       = 0.0    # filter signals below this score
    top_n:              int         = 500


# ═══════════════════════════════════════════════════════════════════════════════
# V3 ENGINE
# ═══════════════════════════════════════════════════════════════════════════════

class SMAOutfitEngineV3:
    """Engine V3 — Level Significance + Sequential Conditional."""

    DEAD_TICKERS_PATH = "/cache/dead_tickers_v3.txt"

    def __init__(
        self,
        client: DataClient,
        cfg: EngineV3Config,
        initial_cache: dict | None = None,
    ):
        self.client       = client
        self.cfg          = cfg
        self.store        = HashMapStore()
        self.system_states: list[SystemState] = []
        self.candle_cache: dict = dict(initial_cache or {})
        self._dead_tickers: set[str] = self._load_dead_tickers()

        # V3 results
        self.level_scores:   list[tuple[HashMapEntry, LevelScore]]   = []
        self.conditions:     list[ConditionalState]                   = []

    def _load_dead_tickers(self) -> set[str]:
        try:
            with open(self.DEAD_TICKERS_PATH) as f:
                return {line.strip() for line in f if line.strip()}
        except FileNotFoundError:
            return set()

    def _save_dead_tickers(self) -> None:
        try:
            os.makedirs(os.path.dirname(self.DEAD_TICKERS_PATH), exist_ok=True)
            with open(self.DEAD_TICKERS_PATH, "w") as f:
                for t in sorted(self._dead_tickers):
                    f.write(t + "\n")
        except Exception as e:
            logging.warning(f"Could not save dead tickers: {e}")

    def _merge_bars(self, existing: pd.DataFrame, new_bars: pd.DataFrame, keep: int = 999) -> pd.DataFrame:
        if existing.empty: return new_bars
        if new_bars.empty: return existing
        combined = pd.concat([existing, new_bars], ignore_index=True)
        combined = combined.drop_duplicates(subset=["timestamp"])
        combined = combined.sort_values("timestamp").reset_index(drop=True)
        return combined.tail(keep).reset_index(drop=True)

    def _prefetch_candles(self) -> None:
        use_incremental = self.cfg.refresh_bars > 0
        cold_pairs, warm_pairs = [], []

        for ticker in self.cfg.universe:
            if ticker in self._dead_tickers:
                continue
            for tf in self.cfg.timeframes:
                key = (ticker, tf)
                if key in self.candle_cache and not self.candle_cache[key].empty:
                    if use_incremental:
                        warm_pairs.append(key)
                else:
                    cold_pairs.append(key)

        total_pairs = len(cold_pairs) + len(warm_pairs)
        if total_pairs == 0:
            return

        max_workers = getattr(self.client, "MAX_WORKERS", 6)
        print(f"  prefetching {len(cold_pairs):,} cold + {len(warm_pairs):,} warm pairs ({max_workers} workers)...", flush=True)

        def fetch_one(args):
            ticker, tf = args
            count = self.cfg.refresh_bars if (use_incremental and (ticker, tf) in self.candle_cache) \
                    else self.cfg.candle_count
            df = self.client.fetch_bars(ticker, tf, count)
            return ticker, tf, df

        all_pairs = cold_pairs + warm_pairs
        ticker_attempts: dict[str, int] = {}
        ticker_failures: dict[str, int] = {}
        for ticker, _ in all_pairs:
            ticker_attempts[ticker] = ticker_attempts.get(ticker, 0) + 1

        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(fetch_one, pair): pair for pair in all_pairs}
            done = 0
            for future in as_completed(futures):
                try:
                    ticker, tf, df = future.result()
                    key = (ticker, tf)
                    if use_incremental and key in self.candle_cache:
                        self.candle_cache[key] = self._merge_bars(
                            self.candle_cache[key], df, keep=self.cfg.candle_count
                        )
                    else:
                        self.candle_cache[key] = df
                except Exception as ex:
                    ticker, tf = futures[future]
                    if (ticker, tf) not in self.candle_cache:
                        self.candle_cache[(ticker, tf)] = pd.DataFrame()
                    ticker_failures[ticker] = ticker_failures.get(ticker, 0) + 1
                done += 1
                if done % 200 == 0:
                    print(f"  prefetched {done:,}/{total_pairs:,}", flush=True)

        new_dead = {t for t, f in ticker_failures.items()
                    if f >= ticker_attempts.get(t, 1)}
        if new_dead:
            self._dead_tickers.update(new_dead)
            self._save_dead_tickers()

    def scan(self) -> None:
        """Full detection pass — identical mechanics to engine 1/2."""
        self._prefetch_candles()

        all_tfs = self.cfg.timeframes
        total = len(self.cfg.universe) * len(all_tfs) * len(self.cfg.outfits)
        print(f"\n  Scanning {total:,} combinations "
              f"({len(self.cfg.universe)} tickers × {len(all_tfs)} tfs × {len(self.cfg.outfits)} outfits)...",
              flush=True)

        n_workers = int(os.environ.get("V3_SCAN_WORKERS",
                        os.environ.get("ENGINE_SCAN_WORKERS", "6")))
        tickers = list(self.cfg.universe)
        chunk_size = math.ceil(len(tickers) / n_workers)
        ticker_chunks = [tickers[i:i + chunk_size] for i in range(0, len(tickers), chunk_size)]

        worker_args = []
        for chunk in ticker_chunks:
            chunk_set = set(chunk)
            cache_slice = {
                (t, tf): df
                for (t, tf), df in self.candle_cache.items()
                if t in chunk_set
            }
            worker_args.append((
                chunk, cache_slice, self.cfg.outfits, all_tfs,
                self.cfg.lookback, self.cfg.hit_mode, self.cfg.hit_tolerance,
            ))

        n_chunks = len(ticker_chunks)
        print(f"  launching {n_chunks} workers...", flush=True)
        scan_start = time.monotonic()

        with multiprocessing.Pool(processes=n_workers, initializer=_scan_worker_init) as pool:
            for i, result in enumerate(pool.imap_unordered(_scan_worker_fn, worker_args), 1):
                elapsed = time.monotonic() - scan_start
                pct = i / n_chunks * 100
                eta = (elapsed / i) * (n_chunks - i) if i < n_chunks else 0
                print(f"  scan {i}/{n_chunks} workers done ({pct:.0f}%) — "
                      f"{elapsed:.0f}s elapsed, ~{eta:.0f}s remaining", flush=True)
                for entry in result.all():
                    k = entry.key
                    if k not in self.store._store:
                        self.store._store[k] = entry
                    else:
                        existing = self.store._store[k]
                        existing.hit_count += entry.hit_count
                        ts_a, ts_b = existing.last_hit_ts, entry.last_hit_ts
                        if ts_a is None:
                            existing.last_hit_ts = ts_b
                        elif ts_b is not None:
                            existing.last_hit_ts = ts_a if ts_a > ts_b else ts_b

        print(f"  merge complete: {len(self.store):,} total combos found", flush=True)

    def monitor_systems(self) -> None:
        self.system_states = evaluate_systems(self.client)

    def rank(self, cumulative_ds: Optional[dict] = None) -> None:
        """Run v3 scoring on all entries in the store."""
        now = pd.Timestamp.now(tz="UTC")

        # Build flat entry list for cross-outfit/cross-tf scoring
        all_entries_flat = [
            {
                "ticker":      e.ticker,
                "timeframe":   e.timeframe,
                "outfit_id":   e.outfit_id,
                "entry_price": e.last_hit_price or 0.0,
            }
            for e in self.store.all()
        ]

        print(f"\n  Scoring {len(self.store):,} entries with v3 composite...", flush=True)

        self.level_scores = rank_entries_v3(
            store           = self.store,
            candle_cache    = self.candle_cache,
            now             = now,
            all_entries_flat = all_entries_flat,
            cumulative_ds   = cumulative_ds,
            lookback        = self.cfg.lookback,
            weight_freshness = self.cfg.weight_freshness,
            min_tf_minutes  = self.cfg.min_tf_minutes,
        )

        print(f"  Scored {len(self.level_scores):,} entries", flush=True)

        # ── Sequential conditional layer ──────────────────────────────────────
        if self.cfg.run_conditions:
            print(f"  Running sequential conditional layer...", flush=True)
            all_store_entries = self.store.all()
            self.conditions = []

            for entry, ls in self.level_scores[:self.cfg.top_n]:
                df = self.candle_cache.get((entry.ticker, entry.timeframe), pd.DataFrame())
                try:
                    cs = evaluate_condition(
                        entry              = entry,
                        df                 = df,
                        all_store_entries  = all_store_entries,
                        candle_cache       = self.candle_cache,
                    )
                    self.conditions.append(cs)
                except Exception as ex:
                    logging.debug(f"Condition eval failed for {entry.ticker}: {ex}")

            summary = summarize_states(self.conditions)
            print(f"  Conditions: STRONG={summary['STRONG']} HOLD={summary['HOLD']} "
                  f"WEAK={summary['WEAK']} IGNORE={summary['IGNORE']}", flush=True)

    def top_n(self, n: int = 500) -> list[dict]:
        """Return top N signals with full v3 data."""
        cond_map = {
            (c.ticker, c.timeframe, c.outfit_id): c
            for c in self.conditions
        }

        out = []
        for rank, (entry, ls) in enumerate(self.level_scores[:n], 1):
            cs = cond_map.get((entry.ticker, entry.timeframe, entry.outfit_id))
            out.append({
                "rank":            rank,
                "ticker":          entry.ticker,
                "timeframe":       entry.timeframe,
                "outfit_id":       entry.outfit_id,
                "outfit_name":     ls.outfit_name,
                "outfit_periods":  list(entry.outfit_periods),
                "hit_count":       entry.hit_count,
                "entry_price":     ls.entry_price,
                "parm_period":     ls.parm_period,
                "parm_price":      ls.parm_price,
                # Composite score
                "composite":       ls.composite,
                "grade":           ls.grade,
                # Score components
                "hit_rate":        round(ls.hit_rate * 100, 1),
                "volume_ratio":    round(ls.volume_ratio, 3),
                "hold_rate":       round(ls.hold_rate * 100, 1),
                "cross_outfits":   ls.outfit_matches,
                "cross_tfs":       ls.tf_matches,
                "recency":         round(ls.recency * 100, 1),
                "persistence":     round(ls.persistence * 100, 1),
                # Conditional state
                "state":           cs.state           if cs else "—",
                "close_vs_parm":   cs.close_vs_parm   if cs else "—",
                "parm_distance":   cs.parm_distance    if cs else 0.0,
                "micro_active":    cs.micro_active     if cs else False,
                "micro_tfs":       cs.micro_tfs        if cs else [],
                "entry":           cs.entry            if cs else ls.entry_price,
                "stop":            cs.stop             if cs else 0.0,
                "target":          cs.target           if cs else 0.0,
                "last_hit_ts":     entry.last_hit_ts.isoformat() if entry.last_hit_ts else None,
            })
        return out


# ═══════════════════════════════════════════════════════════════════════════════
# OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

def save_xlsx_v3(signals: list[dict], path: str) -> None:
    """Save v3 results to Excel with conditional state color coding."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [XLSX] openpyxl not installed — skipping")
        return

    STATE_COLORS = {
        "STRONG": ("1B5E20", "FFFFFF"),
        "HOLD":   ("F57F17", "FFFFFF"),
        "WEAK":   ("1565C0", "FFFFFF"),
        "IGNORE": ("B71C1C", "FFFFFF"),
        "—":      ("F5F5F5", "888888"),
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "V3 Signals"

    thin   = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    headers = [
        "Rank", "State", "Ticker", "Timeframe", "Outfit", "Periods",
        "Score", "Grade", "Hit Count", "Hit Rate%", "Vol Ratio",
        "Hold Rate%", "X-Outfits", "X-TFs", "Recency%", "Persist%",
        "PARM", "PARM Price", "Close vs PARM", "Dist%",
        "Micro TFs", "Entry", "Stop", "Target", "Last Hit"
    ]

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    ws.row_dimensions[1].height = 22
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center
        c.border = border

    for row_idx, sig in enumerate(signals, 2):
        ws.row_dimensions[row_idx].height = 17
        state = sig.get("state", "—")
        bg, fg = STATE_COLORS.get(state, ("FFFFFF", "000000"))
        state_fill = PatternFill("solid", fgColor=bg)
        row_fill   = PatternFill("solid", fgColor="F8F9FA" if row_idx % 2 == 0 else "FFFFFF")

        periods_str = "/".join(str(p) for p in sig.get("outfit_periods", []))
        micro_str   = ",".join(sig.get("micro_tfs", []))

        values = [
            sig["rank"],
            state,
            sig["ticker"],
            sig["timeframe"],
            sig["outfit_name"],
            periods_str,
            sig["composite"],
            sig["grade"],
            sig["hit_count"],
            sig["hit_rate"],
            sig["volume_ratio"],
            sig["hold_rate"],
            sig["cross_outfits"],
            sig["cross_tfs"],
            sig["recency"],
            sig["persistence"],
            sig["parm_period"],
            sig["parm_price"],
            sig["close_vs_parm"],
            sig["parm_distance"],
            micro_str,
            sig["entry"],
            sig["stop"],
            sig["target"],
            sig["last_hit_ts"][:10] if sig["last_hit_ts"] else "—",
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border = border
            c.alignment = center
            if col == 2:  # State column
                c.fill = state_fill
                c.font = Font(bold=True, color=fg, size=10)
            else:
                c.fill = row_fill
                c.font = Font(size=9)

    col_widths = [6,9,8,10,26,32,8,7,10,10,10,11,10,8,10,9,7,11,13,7,10,9,9,9,12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "V3 Engine — Signal Summary"
    ws2["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws2["A1"].fill = PatternFill("solid", fgColor="1F4E79")

    state_counts = {}
    for sig in signals:
        s = sig.get("state", "—")
        state_counts[s] = state_counts.get(s, 0) + 1

    rows = [
        ("Total signals", len(signals)),
        ("STRONG", state_counts.get("STRONG", 0)),
        ("HOLD",   state_counts.get("HOLD",   0)),
        ("WEAK",   state_counts.get("WEAK",   0)),
        ("IGNORE", state_counts.get("IGNORE", 0)),
        ("Actionable (STRONG+HOLD)", state_counts.get("STRONG", 0) + state_counts.get("HOLD", 0)),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for i, (k, v) in enumerate(rows, 3):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws2.cell(row=i, column=2, value=v)
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    wb.save(path)
    print(f"\n  ✅ Saved: {path}", flush=True)


def render_dashboard_v3(signals: list[dict], systems: list[SystemState]) -> str:
    lines = ["═" * 80]
    lines.append("  ENGINE V3 — Level Significance + Sequential Conditional")
    lines.append("═" * 80)

    state_counts = {}
    for sig in signals:
        s = sig.get("state", "—")
        state_counts[s] = state_counts.get(s, 0) + 1

    lines.append(f"  Signals: {len(signals)} total | "
                 f"🟢 STRONG={state_counts.get('STRONG',0)} | "
                 f"🟡 HOLD={state_counts.get('HOLD',0)} | "
                 f"🔵 WEAK={state_counts.get('WEAK',0)} | "
                 f"🔴 IGNORE={state_counts.get('IGNORE',0)}")
    lines.append("─" * 80)

    # Show top 20 actionable
    actionable = [s for s in signals if s["state"] in ("STRONG", "HOLD")][:20]
    if actionable:
        lines.append(f"  {'#':<4} {'St':<8} {'Ticker':<7} {'TF':<6} {'Score':<7} "
                     f"{'Grade':<6} {'PARM':<6} {'vs PARM':<10} {'Micro':<8} "
                     f"{'Entry':>8} {'Stop':>8} {'Target':>8}")
        lines.append("  " + "─" * 76)
        for i, sig in enumerate(actionable, 1):
            micro = ",".join(sig.get("micro_tfs", [])) or "—"
            lines.append(
                f"  {i:<4} {sig['state']:<8} {sig['ticker']:<7} {sig['timeframe']:<6} "
                f"{sig['composite']:<7.1f} {sig['grade']:<6} "
                f"MA{sig['parm_period']:<5} {sig['close_vs_parm']:<10} {micro:<8} "
                f"{sig['entry']:>8.2f} {sig['stop']:>8.2f} {sig['target']:>8.2f}"
            )
    else:
        lines.append("  No STRONG or HOLD signals this cycle.")

    lines.append("─" * 80)
    lines.append("  SYSTEMS:")
    for s in systems:
        glyph = "✅" if s.state == "positive" else ("❌" if s.state == "negative" else "⬜")
        lines.append(f"    {glyph} {s.name:<15} {s.state.upper():<10} {s.note}")
    lines.append("═" * 80)
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# TICKER LOADER
# ═══════════════════════════════════════════════════════════════════════════════

V3_TICKERS_PATH = "/app_v3/v3_tickers.txt"


def load_v3_tickers(path: str = V3_TICKERS_PATH) -> list[str]:
    tickers = []
    try:
        with open(path) as f:
            for line in f:
                line = line.split("#")[0].strip()
                if line:
                    tickers.append(line.upper())
        if tickers:
            print(f"  [TICKERS] ✅ Loaded {len(tickers)} tickers from {path}", flush=True)
    except FileNotFoundError:
        print(f"  [TICKERS] ⚠️  {path} not found — using tier1", flush=True)
    return tickers


def build_universe(tier: str) -> list[str]:
    if tier == "tier1":   raw = UNIVERSE_TIER_1
    elif tier == "tier2": raw = UNIVERSE_TIER_1 + UNIVERSE_TIER_2
    elif tier == "all":   raw = UNIVERSE_TIER_1 + UNIVERSE_TIER_2 + UNIVERSE_TIER_3
    else:                 raw = UNIVERSE_TIER_1
    seen, out = set(), []
    for t in raw:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# INFLUX CHECK
# ═══════════════════════════════════════════════════════════════════════════════

def _check_influx() -> tuple:
    try:
        sys.path.insert(0, "/app")
        from persistence import InfluxPersistence
    except ImportError:
        print("  [INFLUX] ⚠️  persistence.py not found — in-cycle only", flush=True)
        return None, {}

    persist = InfluxPersistence()
    if not persist.enabled or not persist._connected:
        print("  [INFLUX] ❌ NOT CONNECTED — in-cycle scores only", flush=True)
        return persist, {}

    print("  [INFLUX] ✅ Connected — querying cumulative deciseconds...", flush=True)
    try:
        cumulative_ds = persist.query_cumulative_deciseconds(window_days=7)
        if cumulative_ds:
            print(f"  [INFLUX] ✅ {len(cumulative_ds):,} keys loaded (7d window)", flush=True)
        else:
            print("  [INFLUX] ⚠️  Connected but no data — first run", flush=True)
        return persist, cumulative_ds
    except Exception as e:
        print(f"  [INFLUX] ⚠️  Query failed ({e})", flush=True)
        return persist, {}


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(description="SMA Outfit Detection Engine V3")
    parser.add_argument("--source",     choices=["mock", "webull"], default="mock")
    parser.add_argument("--universe",   choices=["tier1", "tier2", "all"], default="tier1")
    parser.add_argument("--tickers",    nargs="+", metavar="TICKER")
    parser.add_argument("--timeframes", nargs="+",
                        default=["1m","5m","15m","30m","1h","2h","4h","1d","1w","1mo"])
    parser.add_argument("--lookback",   type=int, default=390)
    parser.add_argument("--top-n",      type=int, default=500)
    parser.add_argument("--min-score",  type=float, default=0.0,
                        help="Filter signals below this composite score")
    parser.add_argument("--no-conditions", action="store_true",
                        help="Skip sequential conditional layer")
    parser.add_argument("--xlsx",       action="store_true")
    parser.add_argument("--json",       action="store_true")
    parser.add_argument("--verbose",    "-v", action="store_true")
    args = parser.parse_args()

    logging.basicConfig(
        level=logging.INFO if args.verbose else logging.WARNING,
        format="%(asctime)s [%(levelname)s] %(message)s",
    )
    for _noisy in ("webull", "webull.core.client", "webull.core.auth",
                   "urllib3", "urllib3.connectionpool"):
        _lg = logging.getLogger(_noisy)
        _lg.setLevel(logging.CRITICAL)
        _lg.propagate = False

    print("\n" + "═" * 80)
    print("  ENGINE V3 — startup checks")
    print("═" * 80, flush=True)

    persist, cumulative_ds = _check_influx()

    # Webull client
    if args.source == "webull":
        app_key    = os.environ.get("WEBULL_APP_KEY")
        app_secret = os.environ.get("WEBULL_APP_SECRET")
        if not (app_key and app_secret):
            print("  [WEBULL] ❌ WEBULL_APP_KEY/SECRET required", file=sys.stderr)
            sys.exit(1)
        client = WebullClient(app_key, app_secret,
                              region=os.environ.get("WEBULL_REGION", "us"))
        # Force REST — avoids SDK cold-container hang
        client._sdk_available = False
        client._sdk_data_client = None

        # Try SDK with timeout
        import concurrent.futures
        def _try_sdk():
            from webull.core.client import ApiClient
            from webull.data.data_client import DataClient as SDKDataClient
            host = "api.webull.com"
            ac = ApiClient(app_key, app_secret, os.environ.get("WEBULL_REGION","us"))
            ac.add_endpoint(os.environ.get("WEBULL_REGION","us"), host)
            return SDKDataClient(ac)
        try:
            with concurrent.futures.ThreadPoolExecutor(max_workers=1) as ex:
                fut = ex.submit(_try_sdk)
                sdk = fut.result(timeout=30)
            client._sdk_data_client = sdk
            client._sdk_available = True
            mode = "SDK"
        except Exception:
            mode = "REST"
        print(f"  [WEBULL] ✅ Client initialised ({mode})", flush=True)
    else:
        client = MockClient()
        print("  [WEBULL] ⚠️  Mock client", flush=True)

    # Universe
    if args.tickers:
        universe = [t.upper() for t in args.tickers]
        ulabel = f"inline ({len(universe)} tickers)"
    else:
        file_tickers = load_v3_tickers(V3_TICKERS_PATH)
        if file_tickers:
            universe = file_tickers
            ulabel = f"v3_tickers.txt ({len(universe)} tickers)"
        else:
            universe = build_universe(args.universe)
            ulabel = f"{args.universe} ({len(universe)} tickers)"

    seen = set()
    universe = [t for t in universe if not (t in seen or seen.add(t))]

    cfg = EngineV3Config(
        universe       = universe,
        timeframes     = args.timeframes,
        lookback       = args.lookback,
        top_n          = args.top_n,
        run_conditions = not args.no_conditions,
        min_composite  = args.min_score,
    )

    print(f"  [CONFIG] universe={ulabel} tfs={cfg.timeframes} lookback={cfg.lookback}")
    print("═" * 80 + "\n", flush=True)

    engine = SMAOutfitEngineV3(client, cfg)
    engine.monitor_systems()
    engine.scan()
    engine.rank(cumulative_ds=cumulative_ds)

    signals = engine.top_n(args.top_n)

    if args.min_score > 0:
        signals = [s for s in signals if s["composite"] >= args.min_score]

    if persist:
        try:
            persist.flush()
            persist.close()
        except Exception:
            pass

    if args.json:
        print(json.dumps({"signals": signals,
                          "systems": [asdict(s) for s in engine.system_states]},
                         indent=2, default=str))
    else:
        print(render_dashboard_v3(signals, engine.system_states))

    if args.xlsx:
        import os as _os
        _os.makedirs(OUTPUT_DIR, exist_ok=True)
        ts = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M-%S")
        save_xlsx_v3(signals, f"{OUTPUT_DIR}/v3_{ts}.xlsx")


if __name__ == "__main__":
    main()
