#!/usr/bin/env python3
"""
triangulator.py — Real-time cross-engine triangulation terminal UI

Reads latest outputs from:
  - Original engine   (sma_engine/output/signals_current.xlsx)
  - Normalized engine (sma_engine/output/normalized_engine/normalized_*.xlsx)
  - V3 engine         (sma_engine/output/v3/v3_*.xlsx)
  - Discovery         (sma_engine/output/discovery/discovery_*.csv)
  - Confluence        (sma_engine/output/confluence/confluence_*.csv)
  - Trades            (sma_engine/output/trades/trades_*.csv)
  - Backtest          (sma_engine/output/backtest_*.csv)
  - Snapshots         (sma_engine/output/snapshots/snapshot_*.csv)

No modifications to any existing files or folders.

Usage:
    cd ~/Developer/triangulation_engine
    pip install rich watchdog openpyxl
    python3 triangulator.py [--top N] [--refresh S]
"""

from __future__ import annotations

import argparse
import csv
import os
import threading
import time
from dataclasses import dataclass, field
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from rich import box
from rich.console import Console
from rich.layout import Layout
from rich.live import Live
from rich.panel import Panel
from rich.table import Table
from rich.text import Text

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    WATCHDOG_AVAILABLE = True
except ImportError:
    WATCHDOG_AVAILABLE = False


# ── Config ────────────────────────────────────────────────────────────────────

SMA_ENGINE_PATH = Path(os.path.expanduser("~/Developer/sma_engine"))

PATHS = {
    "original":   SMA_ENGINE_PATH / "output" / "signals_current.xlsx",
    "normalized": SMA_ENGINE_PATH / "output" / "normalized_engine",
    "v3":         SMA_ENGINE_PATH / "output" / "v3",
    "discovery":  SMA_ENGINE_PATH / "output" / "discovery",
    "confluence": SMA_ENGINE_PATH / "output" / "confluence",
    "trades":     SMA_ENGINE_PATH / "output" / "trades",
    "backtest":   SMA_ENGINE_PATH / "output",
    "snapshots":  SMA_ENGINE_PATH / "output" / "snapshots",
}

# Macro signal groups — bear ETFs strong = bearish, bull ETFs strong = bullish
MACRO_GROUPS = {
    "S&P 500":     {"bear": ["SPDN","SH","SDS","SPXU","SPXS"],    "bull": ["SPY","SSO","UPRO","SPXL"]},
    "Nasdaq":      {"bear": ["QQQD","PSQ","QID","SQQQ"],           "bull": ["QQQ","QLD","TQQQ","QQQM"]},
    "Dow":         {"bear": ["DOG","DXD","SDOW"],                  "bull": ["DIA","DDM","UDOW"]},
    "Russell":     {"bear": ["RWM","TWM","TZA","SRTY"],            "bull": ["IWM","UWM","TNA","URTY"]},
    "Bonds/Rates": {"bull": ["TLT","TMF","LQD","IEF"],             "bear": ["TBT","TMV","TBF","TBX"]},
    "Gold":        {"bull": ["GLD","IAU","GDX","GDXJ","GLDM","SIL","WPM"], "bear": ["GLL","DGZ","ZSL"]},
    "Commodities": {"bull": ["DBA","PDBC","DBB","USO","UCO","UNG","BOIL"], "bear": ["SCO","KOLD","DUG","DRIP"]},
    "Volatility":  {"bull": ["UVXY","UVIX","VXX","VIXY"],          "bear": ["SVIX","SVXY"]},
}


# ── Data structures ───────────────────────────────────────────────────────────

@dataclass
class TickerSignal:
    ticker: str
    score: float = 0.0

    # Original engine
    orig_rank: int = 0
    orig_conv: str = ""
    orig_tf: str = ""

    # Normalized engine
    norm_rank: int = 0
    norm_conv: str = ""
    norm_tf: str = ""

    # V3 engine
    v3_rank: int = 0
    v3_state: str = ""
    v3_tf: str = ""
    v3_cross_tf: int = 0
    v3_grade: str = ""

    # Discovery
    disc_tf: str = ""
    disc_dir: str = ""
    disc_sma: str = ""

    # Confluence
    conf_score: int = 0

    # Trades
    trade_side: str = ""
    trade_conf: str = ""
    trade_entry: float = 0.0
    trade_stop: float = 0.0
    trade_rr: float = 0.0

    # Backtest
    bt_sharpe: float = 0.0
    bt_winrate: float = 0.0

    # Meta
    engines_confirmed: int = 0


# ── Helpers ───────────────────────────────────────────────────────────────────

def latest_file(directory: Path, pattern: str) -> Optional[Path]:
    files = sorted(directory.glob(pattern)) if directory.exists() else []
    return files[-1] if files else None


def conv_num(conv_str: str) -> int:
    try:
        return int(str(conv_str).split('/')[0])
    except Exception:
        return 0


# ── File readers ──────────────────────────────────────────────────────────────

def read_original() -> dict[str, dict]:
    result: dict[str, dict] = {}
    path = PATHS["original"]
    if not path.exists():
        return result
    try:
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header_idx = next((i for i, r in enumerate(rows) if r[0] == 'Rank'), None)
        if header_idx is None:
            return result
        for row in rows[header_idx + 1:]:
            if not row[0] or not str(row[0]).isdigit():
                continue
            rank, ticker, tf, outfit, hits, conv, score = (row + (None,) * 7)[:7]
            if ticker and str(ticker) not in result:
                result[str(ticker)] = {
                    'rank': int(rank),
                    'tf': str(tf or ''),
                    'conv': str(conv or ''),
                    'score': float(score) if score else 0.0,
                }
    except Exception:
        pass
    return result


def read_normalized() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["normalized"], "normalized_*.xlsx")
    if not f:
        return result
    try:
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row or not row[0]:
                continue
            rank, ticker, tf = row[0], row[1], row[2]
            conv, entry, score = row[7], row[8], row[9]
            if ticker and str(ticker) not in result:
                result[str(ticker)] = {
                    'rank': int(rank) if rank else 0,
                    'tf': str(tf or ''),
                    'conv': str(conv or ''),
                    'score': float(score) if score else 0.0,
                    'entry': float(entry) if entry else 0.0,
                }
    except Exception:
        pass
    return result


def read_v3() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["v3"], "v3_*.xlsx")
    if not f:
        return result
    try:
        wb = openpyxl.load_workbook(f)
        ws = wb['V3 Signals']
        for row in list(ws.iter_rows(values_only=True))[1:]:
            if not row or not row[0]:
                continue
            rank, state, ticker, tf = row[0], row[1], row[2], row[3]
            score, grade = row[6], row[7]
            x_tfs = row[13]
            entry, stop = row[21], row[22]
            if ticker and str(ticker) not in result:
                result[str(ticker)] = {
                    'rank': int(rank) if rank else 0,
                    'state': str(state or ''),
                    'tf': str(tf or ''),
                    'score': float(score) if score else 0.0,
                    'grade': str(grade or ''),
                    'x_tf': int(x_tfs) if x_tfs else 0,
                    'entry': float(entry) if entry else 0.0,
                    'stop': float(stop) if stop else 0.0,
                }
    except Exception:
        pass
    return result


def read_discovery() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["discovery"], "discovery_*.csv")
    if not f:
        return result
    TF_RANK = {'1w':7,'1d':6,'4h':5,'2h':4,'1h':3,'30m':2,'15m':1,'5m':0,'1m':0}
    try:
        with open(f) as fp:
            for row in csv.DictReader(fp):
                t = row['ticker']
                tf = row['timeframe']
                if t not in result or TF_RANK.get(tf, 0) > TF_RANK.get(result[t]['timeframe'], 0):
                    result[t] = row
    except Exception:
        pass
    return result


def read_confluence() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["confluence"], "confluence_*.csv")
    if not f:
        return result
    try:
        with open(f) as fp:
            for row in csv.DictReader(fp):
                result[row['ticker']] = row
    except Exception:
        pass
    return result


def read_trades() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["trades"], "trades_*.csv")
    if not f:
        return result
    try:
        with open(f) as fp:
            for row in csv.DictReader(fp):
                result[row['ticker']] = row
    except Exception:
        pass
    return result


def read_backtest() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["backtest"], "backtest_*.csv")
    if not f:
        return result
    try:
        with open(f) as fp:
            for row in csv.DictReader(fp):
                t = row['ticker']
                sharpe = float(row.get('sharpe', 0) or 0)
                if t not in result or sharpe > float(result[t].get('sharpe', 0) or 0):
                    result[t] = row
    except Exception:
        pass
    return result


def read_snapshot() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["snapshots"], "snapshot_*.csv")
    if not f:
        return result
    try:
        with open(f) as fp:
            for row in csv.DictReader(fp):
                t = row['ticker']
                score = float(row.get('score', 0) or 0)
                if t not in result or score > float(result[t].get('score', 0) or 0):
                    result[t] = row
    except Exception:
        pass
    return result


# ── Triangulation scorer ──────────────────────────────────────────────────────

def triangulate(orig, norm, v3, disc, conf, trades, bt, snap) -> list[TickerSignal]:
    all_tickers = set(orig) | set(norm) | set(v3) | set(disc) | set(conf) | set(trades)

    signals: dict[str, TickerSignal] = {}

    for ticker in all_tickers:
        sig = TickerSignal(ticker=ticker)
        score = 0.0
        engines = 0

        # ── Original engine ──────────────────────────────────────────────────
        if ticker in orig:
            o = orig[ticker]
            sig.orig_rank = o['rank']
            sig.orig_conv = o['conv']
            sig.orig_tf   = o['tf']
            cn = conv_num(o['conv'])
            score += 1.0 + (0.5 if cn >= 2 else 0) + (0.5 if cn >= 3 else 0)
            engines += 1

        # ── Normalized engine ────────────────────────────────────────────────
        if ticker in norm:
            n = norm[ticker]
            sig.norm_rank = n['rank']
            sig.norm_conv = n['conv']
            sig.norm_tf   = n['tf']
            sig.trade_entry = sig.trade_entry or n.get('entry', 0.0)
            cn = conv_num(n['conv'])
            score += 1.0 + (0.5 if cn >= 2 else 0) + (0.5 if cn >= 3 else 0) + (1.0 if cn >= 4 else 0)
            engines += 1

        # ── V3 engine ────────────────────────────────────────────────────────
        if ticker in v3:
            vd = v3[ticker]
            sig.v3_rank     = vd['rank']
            sig.v3_state    = vd['state']
            sig.v3_tf       = vd['tf']
            sig.v3_cross_tf = vd['x_tf']
            sig.v3_grade    = vd['grade']
            sig.trade_entry = sig.trade_entry or vd.get('entry', 0.0)
            state_score = {'STRONG': 2.0, 'HOLD': 1.0, 'WEAK': 0.5, 'IGNORE': 0.0}.get(vd['state'], 0.0)
            score += state_score
            if vd['x_tf'] >= 3: score += 0.5
            if vd['x_tf'] >= 5: score += 0.5
            engines += 1

        # ── Discovery ────────────────────────────────────────────────────────
        if ticker in disc:
            d = disc[ticker]
            sig.disc_tf  = d['timeframe']
            sig.disc_dir = d['direction']
            sig.disc_sma = f"MA{d['sma_period']}"
            TF_RANK = {'1w':7,'1d':6,'4h':5,'2h':4,'1h':3,'30m':2,'15m':1,'5m':0,'1m':0}
            score += 0.5 + TF_RANK.get(d['timeframe'], 0) * 0.1
            try:
                dist_pct = abs(float(d['close']) - float(d['sma_value'])) / float(d['sma_value']) * 100
                if dist_pct < 0.5:
                    score += 0.5
            except Exception:
                pass

        # ── Confluence ───────────────────────────────────────────────────────
        if ticker in conf:
            c = conf[ticker]
            cn = int(c.get('score', 0) or 0)
            sig.conf_score = cn
            score += cn * 0.5

        # ── Trades ───────────────────────────────────────────────────────────
        if ticker in trades:
            tr = trades[ticker]
            sig.trade_side  = tr.get('side', '')
            sig.trade_conf  = tr.get('confidence', '')
            sig.trade_entry = sig.trade_entry or float(tr.get('entry', 0) or 0)
            sig.trade_stop  = float(tr.get('stop', 0) or 0)
            sig.trade_rr    = float(tr.get('rr', 0) or 0)
            score += {'HIGH': 1.0, 'MEDIUM': 0.5}.get(tr.get('confidence', ''), 0.0)

        # ── Backtest ─────────────────────────────────────────────────────────
        if ticker in bt:
            b = bt[ticker]
            sharpe  = float(b.get('sharpe', 0) or 0)
            winrate = float(b.get('win_rate', 0) or 0)
            sig.bt_sharpe  = sharpe
            sig.bt_winrate = winrate
            if winrate > 0.70: score += 0.5
            if sharpe  > 10:   score += 0.5
            if sharpe  > 20:   score += 0.5

        # ── Snapshot bonus ───────────────────────────────────────────────────
        if ticker in snap:
            score += 0.25

        sig.score = round(score, 2)
        sig.engines_confirmed = engines
        signals[ticker] = sig

    return sorted(signals.values(), key=lambda s: -s.score)


# ── Macro analysis ────────────────────────────────────────────────────────────

def analyze_macro(signals: list[TickerSignal]) -> dict[str, tuple[str, float, float]]:
    """Returns {group: (direction, bear_score, bull_score)}"""
    by_ticker = {s.ticker: s.score for s in signals}
    macro: dict[str, tuple[str, float, float]] = {}

    for group, sides in MACRO_GROUPS.items():
        bear = sum(by_ticker.get(t, 0.0) for t in sides.get('bear', []))
        bull = sum(by_ticker.get(t, 0.0) for t in sides.get('bull', []))

        if bear == 0 and bull == 0:
            direction = "neutral"
        elif bear > 0 and bull == 0:
            direction = "bearish"
        elif bull > 0 and bear == 0:
            direction = "bullish"
        elif bear > bull * 1.3:
            direction = "bearish"
        elif bull > bear * 1.3:
            direction = "bullish"
        else:
            direction = "mixed"

        macro[group] = (direction, bear, bull)

    return macro


# ── Engine status ─────────────────────────────────────────────────────────────

def get_engine_status() -> dict[str, str]:
    status: dict[str, str] = {}
    checks = {
        "original":   (PATHS["original"], None),
        "normalized": (PATHS["normalized"], "normalized_*.xlsx"),
        "v3":         (PATHS["v3"],         "v3_*.xlsx"),
    }
    now = time.time()
    for name, (path, pattern) in checks.items():
        if pattern:
            f = latest_file(path, pattern)
        else:
            f = path if path.exists() else None

        if not f or not f.exists():
            status[name] = "[dim]no file[/dim]"
            continue

        age = now - f.stat().st_mtime
        mtime = datetime.fromtimestamp(f.stat().st_mtime).strftime("%H:%M:%S")
        if age < 900:       # < 15 min
            status[name] = f"[green]{mtime}[/green]"
        elif age < 3600:    # < 1 hr
            status[name] = f"[yellow]{mtime}[/yellow]"
        else:
            status[name] = f"[dim]{mtime}[/dim]"

    return status


# ── Rich UI builders ──────────────────────────────────────────────────────────

DIR_STYLE = {"bearish": "red", "bullish": "green", "mixed": "yellow", "neutral": "dim"}
DIR_ICON  = {"bearish": "▼", "bullish": "▲", "mixed": "◆", "neutral": "─"}
STATE_STYLE = {"STRONG": "bright_green", "HOLD": "yellow", "WEAK": "red", "IGNORE": "dim"}
CONV_STYLE  = {"4/4": "bright_green", "3/4": "green", "2/4": "yellow", "1/4": "dim"}


def build_header(now: str, engine_status: dict[str, str]) -> Panel:
    orig = engine_status.get("original", "—")
    norm = engine_status.get("normalized", "—")
    v3   = engine_status.get("v3", "—")
    subtitle = f"orig {orig}  norm {norm}  v3 {v3}"
    return Panel(
        Text(
            f"TRIANGULATION ENGINE  ✦  {now}",
            style="bold white",
            justify="center",
        ),
        subtitle=subtitle,
        border_style="bright_white",
        padding=(0, 1),
    )


def build_macro_panel(macro: dict[str, tuple[str, float, float]]) -> Panel:
    table = Table(box=None, padding=(0, 3), show_header=False, expand=True)
    table.add_column("Group", style="bold", width=16)
    table.add_column("Direction", width=14)
    table.add_column("Bear score", justify="right", width=12)
    table.add_column("Bull score", justify="right", width=12)

    for group, (direction, bear, bull) in macro.items():
        icon  = DIR_ICON.get(direction, "")
        style = DIR_STYLE.get(direction, "white")
        table.add_row(
            group,
            Text(f"{icon}  {direction.upper()}", style=style),
            Text(f"{bear:.1f}", style="red"   if bear > 0 else "dim"),
            Text(f"{bull:.1f}", style="green" if bull > 0 else "dim"),
        )

    return Panel(table, title="[bold cyan]MACRO[/bold cyan]", border_style="cyan")


def build_signals_table(signals: list[TickerSignal], top_n: int) -> Panel:
    table = Table(
        box=box.SIMPLE_HEAD,
        padding=(0, 1),
        show_header=True,
        header_style="bold white",
        expand=True,
    )
    table.add_column("#",      width=4,  style="dim")
    table.add_column("Ticker", width=8,  style="bold")
    table.add_column("Score",  width=7)
    table.add_column("Eng",    width=5)
    table.add_column("V3",     width=8)
    table.add_column("Norm",   width=6)
    table.add_column("Orig",   width=6)
    table.add_column("Disc",   width=11)
    table.add_column("Trade",  width=9)
    table.add_column("Sharpe", width=7)
    table.add_column("Entry",  width=8)

    shown = [s for s in signals if s.score > 0][:top_n]

    for i, sig in enumerate(shown, 1):
        # Score colour
        if   sig.score >= 7: sc = "bright_green bold"
        elif sig.score >= 5: sc = "green"
        elif sig.score >= 3: sc = "yellow"
        else:                sc = "dim"

        # Engines
        eng_str   = f"{sig.engines_confirmed}/3"
        eng_style = "bright_green" if sig.engines_confirmed == 3 else (
                    "green"        if sig.engines_confirmed == 2 else "dim")

        # V3
        v3_str   = sig.v3_state if sig.v3_state else "—"
        v3_style = STATE_STYLE.get(sig.v3_state, "dim")

        # Norm
        norm_str   = sig.norm_conv if sig.norm_conv else "—"
        norm_style = CONV_STYLE.get(sig.norm_conv, "dim")

        # Orig
        orig_str = sig.orig_conv if sig.orig_conv else "—"

        # Discovery
        if sig.disc_tf:
            disc_str   = f"{sig.disc_tf} {sig.disc_sma}"
            disc_style = "bright_green" if sig.disc_tf in ("1w","1d") else (
                         "green"        if sig.disc_tf in ("4h","2h") else "yellow")
        else:
            disc_str, disc_style = "—", "dim"

        # Trade
        if sig.trade_conf == "HIGH":
            trade_str   = f"{'↑' if sig.trade_side=='BUY' else '↓'} HIGH"
            trade_style = "bright_green" if sig.trade_side == "BUY" else "red"
        elif sig.trade_conf:
            trade_str, trade_style = sig.trade_conf, "yellow"
        else:
            trade_str, trade_style = "—", "dim"

        # Sharpe
        sh_str   = f"{sig.bt_sharpe:.1f}" if sig.bt_sharpe else "—"
        sh_style = "bright_green" if sig.bt_sharpe > 15 else (
                   "green"        if sig.bt_sharpe > 10 else "dim")

        entry_str = f"${sig.trade_entry:.2f}" if sig.trade_entry else "—"

        table.add_row(
            str(i),
            sig.ticker,
            Text(f"{sig.score:.1f}", style=sc),
            Text(eng_str,   style=eng_style),
            Text(v3_str,    style=v3_style),
            Text(norm_str,  style=norm_style),
            orig_str,
            Text(disc_str,  style=disc_style),
            Text(trade_str, style=trade_style),
            Text(sh_str,    style=sh_style),
            entry_str,
        )

    return Panel(
        table,
        title=f"[bold yellow]TRIANGULATED SIGNALS — TOP {top_n}[/bold yellow]",
        subtitle=f"[dim]{len(shown)} shown · {len(signals)} total tickers[/dim]",
        border_style="yellow",
    )


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Triangulation Engine — real-time terminal UI")
    parser.add_argument("--top",     type=int, default=25,  help="Signals to display (default: 25)")
    parser.add_argument("--refresh", type=int, default=30,  help="Refresh interval in seconds (default: 30)")
    args = parser.parse_args()

    console = Console()

    layout = Layout()
    layout.split_column(
        Layout(name="header",  size=3),
        Layout(name="macro",   size=12),
        Layout(name="signals"),
    )

    refresh_event = threading.Event()

    # ── File watcher ──────────────────────────────────────────────────────────
    observer = None
    if WATCHDOG_AVAILABLE:
        class ChangeHandler(FileSystemEventHandler):
            def on_created(self, event):
                if not event.is_directory and any(
                    event.src_path.endswith(ext) for ext in ('.xlsx', '.csv')
                ):
                    refresh_event.set()

        observer = Observer()
        watch_dirs = [PATHS["normalized"], PATHS["v3"], PATHS["discovery"], PATHS["confluence"]]
        for wd in watch_dirs:
            if wd.exists():
                observer.schedule(ChangeHandler(), str(wd), recursive=False)
        observer.start()

    # ── Live UI loop ──────────────────────────────────────────────────────────
    try:
        with Live(layout, console=console, refresh_per_second=2, screen=True):
            while True:
                # Load all data
                orig  = read_original()
                norm  = read_normalized()
                v3    = read_v3()
                disc  = read_discovery()
                conf  = read_confluence()
                tr    = read_trades()
                bt    = read_backtest()
                snap  = read_snapshot()

                signals = triangulate(orig, norm, v3, disc, conf, tr, bt, snap)
                macro   = analyze_macro(signals)
                status  = get_engine_status()
                now     = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

                layout["header"].update(build_header(now, status))
                layout["macro"].update(build_macro_panel(macro))
                layout["signals"].update(build_signals_table(signals, args.top))

                # Wait for file change or interval
                refresh_event.wait(timeout=args.refresh)
                refresh_event.clear()

    except KeyboardInterrupt:
        pass
    finally:
        if observer:
            observer.stop()
            observer.join()
        console.print("\n[dim]Triangulation engine stopped.[/dim]")


if __name__ == "__main__":
    main()
