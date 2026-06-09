"""
heatmap_normalized.py — Heatmap dashboard for the Normalized Engine output.

Reads all normalized engine xlsx files from output/normalized_engine/ and
produces a heatmap Excel file with two sheets:

  Sheet 1 — CROSS-TIMEFRAME (works with a single run)
    Rows    = top tickers by overall normalized score
    Columns = each timeframe scanned
    Cell    = rank on that timeframe (blank if not ranked)
    Color   = green gradient (dark = top rank, light = lower rank)
    Tells you: which tickers have strong signals across multiple timeframes

  Sheet 2 — RISING TICKERS (populates as more runs accumulate)
    Rows    = tickers that appeared in any run
    Columns = each run (timestamped)
    Cell    = overall rank in that run
    Color   = green if rank improved vs prior run, red if declined
    Tells you: which tickers are consistently climbing the leaderboard

Usage:
    docker exec e47_engine python /app/heatmap_normalized.py
    docker exec e47_engine python /app/heatmap_normalized.py --top 100
    docker exec e47_engine python /app/heatmap_normalized.py --top 200 --runs 10
"""

from __future__ import annotations

import argparse
import glob
import os
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    raise

try:
    import pandas as pd
except ImportError:
    print("ERROR: pandas not installed.")
    raise


# ── Directories ───────────────────────────────────────────────────────────────
INPUT_DIR  = Path("/cache/output/normalized_engine")
OUTPUT_DIR = Path("/cache/output/normalized_engine")

# ── Timeframe order (left → right on cross-TF sheet) ─────────────────────────
TF_ORDER = ["1m", "5m", "15m", "30m", "1h", "2h", "4h", "1d", "1w", "1mo"]

# ── Rank-to-color gradient (green scale, dark = best rank) ───────────────────
def rank_color(rank: int | None, max_rank: int) -> str:
    """Return a hex fill color based on rank. None = white (not present)."""
    if rank is None:
        return "FFFFFF"
    pct = 1.0 - min(rank - 1, max_rank - 1) / max_rank  # 1.0 = rank 1, 0.0 = max
    # Interpolate from light green (E8F5E9) to dark green (1B5E20)
    light = (232, 245, 233)
    dark  = (27,  94,  32)
    r = int(light[0] + (dark[0] - light[0]) * pct)
    g = int(light[1] + (dark[1] - light[1]) * pct)
    b = int(light[2] + (dark[2] - light[2]) * pct)
    return f"{r:02X}{g:02X}{b:02X}"


def rank_font_color(rank: int | None, max_rank: int) -> str:
    """White text for dark cells, black for light cells."""
    if rank is None:
        return "AAAAAA"
    pct = 1.0 - min(rank - 1, max_rank - 1) / max_rank
    return "FFFFFF" if pct > 0.5 else "1A1A1A"


def trend_color(rank: int | None, prev_rank: int | None) -> str:
    """
    Green if rank improved (lower number), red if declined, yellow if same,
    light grey if new entry, white if absent.
    """
    if rank is None:
        return "F5F5F5"
    if prev_rank is None:
        return "C8E6C9"   # new entry — light green
    diff = prev_rank - rank  # positive = improved
    if diff > 50:  return "1B5E20"   # big jump up  — dark green
    if diff > 10:  return "43A047"   # solid up     — medium green
    if diff > 0:   return "A5D6A7"   # slight up    — light green
    if diff == 0:  return "FFF9C4"   # flat         — yellow
    if diff > -10: return "EF9A9A"   # slight down  — light red
    if diff > -50: return "E53935"   # solid down   — red
    return "B71C1C"                  # big drop     — dark red


def load_runs(input_dir: Path, max_runs: int = 20) -> list[tuple[str, pd.DataFrame]]:
    """
    Load the most recent `max_runs` xlsx files from the normalized engine output dir.
    Returns list of (timestamp_label, DataFrame) sorted oldest → newest.
    """
    files = sorted(glob.glob(str(input_dir / "normalized_*.xlsx")))
    files = files[-max_runs:]  # keep most recent N

    runs = []
    for f in files:
        try:
            df = pd.read_excel(f, engine="openpyxl")
            # Normalise column names (strip whitespace)
            df.columns = [c.strip() for c in df.columns]
            # Extract timestamp from filename: normalized_2026-06-04_22-00-00.xlsx
            stem = Path(f).stem  # normalized_2026-06-04_22-00-00
            ts_part = stem.replace("normalized_", "")  # 2026-06-04_22-00-00
            ts_label = ts_part.replace("_", " ")[:16]  # 2026-06-04 22:00
            runs.append((ts_label, df))
        except Exception as e:
            print(f"  Warning: could not load {f}: {e}")

    return runs


# ── Sheet 1: Cross-Timeframe Heatmap ─────────────────────────────────────────

def build_cross_tf_sheet(ws, df: pd.DataFrame, top: int) -> None:
    """
    Build the cross-timeframe heatmap from a single run's DataFrame.
    Rows = top tickers, Columns = timeframes.
    """
    # ── Header styling ────────────────────────────────────────────────────────
    title_font   = Font(bold=True, size=13, color="FFFFFF")
    header_fill  = PatternFill("solid", fgColor="1A237E")
    header_font  = Font(bold=True, color="FFFFFF", size=10)
    center       = Alignment(horizontal="center", vertical="center")
    tf_fill      = PatternFill("solid", fgColor="283593")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    # Title
    ws.merge_cells("A1:L1")
    title_cell = ws["A1"]
    title_cell.value = "NORMALIZED ENGINE — Cross-Timeframe Signal Heatmap"
    title_cell.font  = title_font
    title_cell.fill  = header_fill
    title_cell.alignment = center

    # Column headers: Ticker | Overall Rank | Score | TF1 | TF2 | ...
    headers = ["Ticker", "Best Rank", "Norm Score"] + TF_ORDER
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.fill = tf_fill if col > 3 else header_fill
        cell.font = header_font
        cell.alignment = center

    # ── Build ticker → {tf: rank} mapping ────────────────────────────────────
    # Get unique tickers ranked by best overall rank
    ticker_best: dict[str, dict] = {}
    for _, row in df.iterrows():
        ticker = str(row.get("Ticker", "")).strip()
        if not ticker:
            continue
        rank  = row.get("Rank")
        score = row.get("Norm Score")
        tf    = str(row.get("Timeframe", "")).strip()
        if ticker not in ticker_best:
            ticker_best[ticker] = {"best_rank": rank, "score": score, "tf_ranks": {}}
        if rank < ticker_best[ticker]["best_rank"]:
            ticker_best[ticker]["best_rank"] = rank
            ticker_best[ticker]["score"]     = score
        ticker_best[ticker]["tf_ranks"][tf] = rank

    # Sort by best rank, take top N
    sorted_tickers = sorted(ticker_best.items(), key=lambda x: x[1]["best_rank"])[:top]

    # ── Fill rows ─────────────────────────────────────────────────────────────
    max_rank = len(df)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, (ticker, data) in enumerate(sorted_tickers, 3):
        ws.row_dimensions[row_idx].height = 18

        # Ticker
        c = ws.cell(row=row_idx, column=1, value=ticker)
        c.font = Font(bold=True, size=10)
        c.alignment = center
        c.border = border

        # Best rank
        c = ws.cell(row=row_idx, column=2, value=data["best_rank"])
        c.alignment = center
        c.border = border

        # Score
        c = ws.cell(row=row_idx, column=3, value=round(float(data["score"] or 0), 3))
        c.alignment = center
        c.border = border

        # Timeframe rank cells
        for tf_idx, tf in enumerate(TF_ORDER, 4):
            tf_rank = data["tf_ranks"].get(tf)
            hex_fill  = rank_color(tf_rank, max_rank)
            font_col  = rank_font_color(tf_rank, max_rank)
            c = ws.cell(row=row_idx, column=tf_idx,
                        value=tf_rank if tf_rank is not None else "")
            c.fill      = PatternFill("solid", fgColor=hex_fill)
            c.font      = Font(size=9, color=font_col)
            c.alignment = center
            c.border    = border

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 12
    for i, tf in enumerate(TF_ORDER, 4):
        ws.column_dimensions[get_column_letter(i)].width = 7

    ws.freeze_panes = "A3"

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_row = len(sorted_tickers) + 5
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    legend = [
        ("1B5E20", "FFFFFF", "Rank 1–10"),
        ("43A047", "FFFFFF", "Rank 11–50"),
        ("A5D6A7", "1A1A1A", "Rank 51–200"),
        ("C8E6C9", "1A1A1A", "Rank 201–500"),
        ("E8F5E9", "AAAAAA", "Rank 500+"),
        ("FFFFFF", "AAAAAA", "Not ranked"),
    ]
    for i, (bg, fg, label) in enumerate(legend, 2):
        c = ws.cell(row=legend_row, column=i, value=label)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(size=9, color=fg)
        c.alignment = center
        c.border = border


# ── Sheet 2: Rising Tickers Heatmap ──────────────────────────────────────────

def build_rising_sheet(ws, runs: list[tuple[str, pd.DataFrame]], top: int) -> None:
    """
    Build the rising tickers heatmap across multiple runs.
    Rows = tickers, Columns = run timestamps.
    """
    header_fill = PatternFill("solid", fgColor="4A148C")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Side(style="thin", color="DDDDDD")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22

    # Title
    end_col = get_column_letter(2 + len(runs))
    ws.merge_cells(f"A1:{end_col}1")
    title_cell = ws["A1"]
    title_cell.value = "NORMALIZED ENGINE — Rising Tickers (rank trajectory across runs)"
    title_cell.font  = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill  = header_fill
    title_cell.alignment = center

    # Column headers
    ws.cell(row=2, column=1, value="Ticker").font  = header_font
    ws.cell(row=2, column=1).fill = header_fill
    ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=2, value="Latest Rank").font = header_font
    ws.cell(row=2, column=2).fill = header_fill
    ws.cell(row=2, column=2).alignment = center

    for col, (ts_label, _) in enumerate(runs, 3):
        c = ws.cell(row=2, column=col, value=ts_label)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center

    # ── Build ticker rank history ─────────────────────────────────────────────
    # {ticker: [rank_run0, rank_run1, ...]} — None if not in that run
    ticker_history: dict[str, list[int | None]] = {}
    for ts_label, df in runs:
        present = set()
        for _, row in df.iterrows():
            ticker = str(row.get("Ticker", "")).strip()
            rank   = row.get("Rank")
            if ticker and rank is not None:
                if ticker not in ticker_history:
                    ticker_history[ticker] = [None] * len(runs)
                run_idx = [r[0] for r in runs].index(ts_label)
                ticker_history[ticker][run_idx] = int(rank)
                present.add(ticker)

    # Sort by latest run rank (ascending), then by number of appearances
    def sort_key(item):
        ranks = item[1]
        latest = next((r for r in reversed(ranks) if r is not None), 99999)
        appearances = sum(1 for r in ranks if r is not None)
        return (latest, -appearances)

    sorted_tickers = sorted(ticker_history.items(), key=sort_key)[:top]

    # ── Fill rows ─────────────────────────────────────────────────────────────
    for row_idx, (ticker, rank_history) in enumerate(sorted_tickers, 3):
        ws.row_dimensions[row_idx].height = 18

        latest_rank = next((r for r in reversed(rank_history) if r is not None), None)

        c = ws.cell(row=row_idx, column=1, value=ticker)
        c.font = Font(bold=True, size=10)
        c.alignment = center
        c.border = border

        c = ws.cell(row=row_idx, column=2, value=latest_rank)
        c.alignment = center
        c.border = border

        prev_rank = None
        for col, rank in enumerate(rank_history, 3):
            hex_fill = trend_color(rank, prev_rank)
            font_col = "FFFFFF" if hex_fill in ("1B5E20", "4A148C", "B71C1C", "E53935") else "1A1A1A"
            c = ws.cell(row=row_idx, column=col,
                        value=rank if rank is not None else "")
            c.fill      = PatternFill("solid", fgColor=hex_fill)
            c.font      = Font(size=9, color=font_col)
            c.alignment = center
            c.border    = border
            if rank is not None:
                prev_rank = rank

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    for col in range(3, 3 + len(runs)):
        ws.column_dimensions[get_column_letter(col)].width = 16

    ws.freeze_panes = "C3"

    # ── Legend ────────────────────────────────────────────────────────────────
    legend_row = len(sorted_tickers) + 5
    ws.cell(row=legend_row, column=1, value="Legend:").font = Font(bold=True)
    legend = [
        ("1B5E20", "FFFFFF", "Big jump ↑↑"),
        ("43A047", "FFFFFF", "Rising ↑"),
        ("A5D6A7", "1A1A1A", "Slight ↑"),
        ("FFF9C4", "1A1A1A", "Flat →"),
        ("EF9A9A", "1A1A1A", "Slight ↓"),
        ("E53935", "FFFFFF", "Falling ↓"),
        ("B71C1C", "FFFFFF", "Big drop ↓↓"),
        ("C8E6C9", "1A1A1A", "New entry"),
    ]
    for i, (bg, fg, label) in enumerate(legend, 2):
        c = ws.cell(row=legend_row, column=i, value=label)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(size=9, color=fg)
        c.alignment = center
        c.border = Border(
            left=Side(style="thin", color="DDDDDD"),
            right=Side(style="thin", color="DDDDDD"),
            top=Side(style="thin", color="DDDDDD"),
            bottom=Side(style="thin", color="DDDDDD"),
        )


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Normalized Engine Heatmap Dashboard")
    parser.add_argument("--top",  type=int, default=150, help="Top N tickers to show (default: 150)")
    parser.add_argument("--runs", type=int, default=20,  help="Max recent runs to include (default: 20)")
    args = parser.parse_args()

    print("\n" + "═" * 71)
    print("  HEATMAP GENERATOR — Normalized Engine")
    print("═" * 71)

    if not INPUT_DIR.exists():
        print(f"  ERROR: No output found at {INPUT_DIR}")
        print("  Run the normalized engine first:")
        print("    docker exec e47_engine python /app/engine_normalized.py --source webull --universe all --top-n 2000 --xlsx")
        return

    runs = load_runs(INPUT_DIR, max_runs=args.runs)
    if not runs:
        print(f"  ERROR: No normalized_*.xlsx files found in {INPUT_DIR}")
        return

    print(f"  Loaded {len(runs)} run(s)")
    latest_ts, latest_df = runs[-1]
    print(f"  Latest run: {latest_ts} ({len(latest_df)} signals)")

    wb = openpyxl.Workbook()

    # Sheet 1 — Cross-Timeframe
    ws1 = wb.active
    ws1.title = "Cross-Timeframe"
    print(f"  Building cross-timeframe heatmap (top {args.top} tickers)...")
    build_cross_tf_sheet(ws1, latest_df, top=args.top)

    # Sheet 2 — Rising Tickers
    ws2 = wb.create_sheet("Rising Tickers")
    print(f"  Building rising tickers heatmap ({len(runs)} run(s))...")
    build_rising_sheet(ws2, runs, top=args.top)

    # Save
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M-%S")
    out_path = OUTPUT_DIR / f"heatmap_{ts}.xlsx"
    wb.save(str(out_path))

    print(f"\n  ✅ Saved: {out_path}")
    print("═" * 71 + "\n")


if __name__ == "__main__":
    main()
