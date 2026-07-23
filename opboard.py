#!/usr/bin/env python3
"""
opboard.py — Operator dashboard: raw combined view per ticker across all engines.

Unlike the triangulator (which scores and ranks), this shows RAW data from every
source side by side — one row per ticker, no opinion. Read the room yourself.

Sources (read-only, no files modified):
  - Main engine       output/signals_current.xlsx
  - Normalized engine output/normalized_engine/normalized_*.xlsx
  - V3 engine         output/v3/v3_*.xlsx
  - Discovery         output/discovery/discovery_*.csv
  - Confluence        output/confluence/confluence_*.csv
  - Trades            output/trades/trades_*.csv
  - Backtest          output/backtest_*.csv
  - OHLC              output/ohlc_log.csv (latest row per ticker)
  - Cumulative DS     InfluxDB localhost:8086 (7-day deciseconds per level,
                      queried for displayed tickers only, cached 5 min)

Usage:
    cd ~/Developer/sma_engine && python3 opboard.py [--top N] [--refresh S]
                                                    [--min-src N] [--ticker SYM]
                                                    [--sort src|cumds]

    --top      rows to display (default 40)
    --refresh  seconds between refreshes (default 30)
    --min-src  only show tickers present in at least N sources (default 1)
    --ticker   show only this ticker (case-insensitive)
    --sort     src   = most sources first (default)
               cumds = most cumulative deciseconds at best level first

Ctrl+C to exit. This is a viewer — do NOT add it to eod.sh or fullrun.sh.
"""

from __future__ import annotations

import argparse
import csv
import os
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

import openpyxl
from rich import box
from rich.console import Console
from rich.layout import Layout
from rich.live import Live
from rich.panel import Panel
from rich.table import Table
from rich.text import Text


# ── Config ────────────────────────────────────────────────────────────────────

BASE = Path(os.environ.get('OPBOARD_BASE', os.path.expanduser('~/Developer/sma_engine')))

PATHS = {
    "main":       BASE / "output" / "signals_current.xlsx",
    "normalized": BASE / "output" / "normalized_engine",
    "v3":         BASE / "output" / "v3",
    "discovery":  BASE / "output" / "discovery",
    "confluence": BASE / "output" / "confluence",
    "trades":     BASE / "output" / "trades",
    "backtest":   BASE / "output",
}

SOURCE_LABELS = ["main", "norm", "v3", "disc", "conf", "trade", "bt"]


# ── Row model ─────────────────────────────────────────────────────────────────

@dataclass
class OperatorRow:
    ticker: str
    sources: int = 0

    # Main engine
    main_rank: int = 0
    main_tf: str = ""
    main_outfit: str = ""
    main_hits: int = 0
    main_conv: str = ""

    # Normalized
    norm_rank: int = 0
    norm_tf: str = ""
    norm_conv: str = ""
    norm_entry: float = 0.0

    # V3
    v3_rank: int = 0
    v3_state: str = ""
    v3_tf: str = ""
    v3_grade: str = ""
    v3_xtf: int = 0

    # Discovery
    disc_tf: str = ""
    disc_sma: str = ""
    disc_dir: str = ""
    disc_sma_value: float = 0.0
    disc_close: float = 0.0
    disc_high: float = 0.0
    disc_low: float = 0.0

    # Confluence
    conf_score: int = 0

    # Trades
    trade_side: str = ""
    trade_conf: str = ""
    trade_entry: float = 0.0
    trade_stop: float = 0.0
    trade_rr: float = 0.0

    # Backtest
    bt_sharpe: float = 0.0
    bt_winrate: float = 0.0

    # OHLC (latest row per ticker from ohlc_log.csv — candle of the signal's
    # timeframe, NOT the daily session; ohlc_tf says which)
    ohlc_o: float = 0.0
    ohlc_h: float = 0.0
    ohlc_l: float = 0.0
    ohlc_c: float = 0.0
    ohlc_tf: str = ""

    # Cumulative deciseconds — best (max) level for this ticker, 7-day window
    cum_ds: float = 0.0

    # 7-day appearance counts per engine (main cycles, norm cycles, v3 STRONG/HOLD)
    h7_main: int = 0
    h7_norm: int = 0
    h7_v3: int = 0

    @property
    def price(self) -> float:
        """Best available price: OHLC close > discovery close > trade entry > norm entry."""
        return self.ohlc_c or self.disc_close or self.trade_entry or self.norm_entry or 0.0

    @property
    def entry_ref(self) -> float:
        """Best entry/level reference: trade entry > norm entry > discovery SMA value."""
        return self.trade_entry or self.norm_entry or self.disc_sma_value or 0.0

    @property
    def long_short(self) -> str:
        """'L' if price is above the entry reference, 'S' if below, '' if unknown."""
        if not self.price or not self.entry_ref:
            return ""
        return "L" if self.price >= self.entry_ref else "S"


# ── Helpers ───────────────────────────────────────────────────────────────────

def latest_file(directory: Path, pattern: str) -> Optional[Path]:
    files = sorted(directory.glob(pattern)) if directory.exists() else []
    return files[-1] if files else None


# ── Readers (same parsing as triangulator.py, kept read-only) ─────────────────

def read_main() -> dict[str, dict]:
    result: dict[str, dict] = {}
    path = PATHS["main"]
    if not path.exists():
        return result
    try:
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb["Current"] if "Current" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        header_idx = next((i for i, r in enumerate(rows) if r and r[0] == "Rank"), None)
        if header_idx is None:
            return result
        for row in rows[header_idx + 1:]:
            if not row or row[0] is None or not str(row[0]).isdigit():
                continue
            rank, ticker, tf, outfit, hits, conv, score = (tuple(row) + (None,) * 7)[:7]
            t = str(ticker or "")
            if t and t not in result:
                result[t] = {
                    "rank": int(rank), "tf": str(tf or ""), "outfit": str(outfit or ""),
                    "hits": int(hits or 0), "conv": str(conv or ""),
                }
    except Exception:
        pass
    return result


def read_normalized() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["normalized"], "normalized_*.xlsx")
    if not f:
        return result
    try:
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            rank, ticker, tf = row[0], row[1], row[2]
            conv = row[7] if len(row) > 7 else None
            entry = row[8] if len(row) > 8 else None
            t = str(ticker or "")
            if t and t not in result:
                result[t] = {
                    "rank": int(rank) if rank else 0, "tf": str(tf or ""),
                    "conv": str(conv or ""),
                    "entry": float(entry) if entry else 0.0,
                }
    except Exception:
        pass
    return result


def read_v3() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["v3"], "v3_*.xlsx")
    if not f:
        return result
    try:
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb["V3 Signals"] if "V3 Signals" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        for row in rows[1:]:
            if not row or not row[0]:
                continue
            rank, state, ticker, tf = row[0], row[1], row[2], row[3]
            grade = row[7] if len(row) > 7 else None
            x_tfs = row[13] if len(row) > 13 else None
            t = str(ticker or "")
            if t and t not in result:
                result[t] = {
                    "rank": int(rank) if rank else 0, "state": str(state or ""),
                    "tf": str(tf or ""), "grade": str(grade or ""),
                    "x_tf": int(x_tfs) if x_tfs else 0,
                }
    except Exception:
        pass
    return result


def read_discovery() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["discovery"], "discovery_*.csv")
    if not f:
        return result
    TF_RANK = {"1mo": 8, "1w": 7, "1d": 6, "4h": 5, "2h": 4, "1h": 3, "30m": 2, "15m": 1}
    try:
        with open(f) as fp:
            for row in csv.DictReader(fp):
                t = row["ticker"]
                tf = row["timeframe"]
                if t not in result or TF_RANK.get(tf, 0) > TF_RANK.get(result[t]["timeframe"], 0):
                    result[t] = row
    except Exception:
        pass
    return result


def read_confluence() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["confluence"], "confluence_*.csv")
    if not f:
        return result
    try:
        with open(f) as fp:
            for row in csv.DictReader(fp):
                result[row["ticker"]] = row
    except Exception:
        pass
    return result


def read_trades() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["trades"], "trades_*.csv")
    if not f:
        return result
    try:
        with open(f) as fp:
            for row in csv.DictReader(fp):
                result[row["ticker"]] = row
    except Exception:
        pass
    return result


def read_backtest() -> dict[str, dict]:
    result: dict[str, dict] = {}
    f = latest_file(PATHS["backtest"], "backtest_*.csv")
    if not f:
        return result
    try:
        with open(f) as fp:
            for row in csv.DictReader(fp):
                t = row["ticker"]
                sharpe = float(row.get("sharpe", 0) or 0)
                if t not in result or sharpe > float(result[t].get("sharpe", 0) or 0):
                    result[t] = row
    except Exception:
        pass
    return result


def read_ohlc() -> dict[str, dict]:
    """Latest OHLC per ticker from ohlc_log.csv (main engine appends each cycle)."""
    result: dict[str, dict] = {}
    path = BASE / "output" / "ohlc_log.csv"
    if not path.exists():
        return result
    try:
        with open(path) as fp:
            for row in csv.DictReader(fp):
                result[row["ticker"]] = row   # later rows overwrite = latest wins
    except Exception:
        pass
    return result


# ── Cumulative deciseconds (InfluxDB, cached) ─────────────────────────────────

_CUMDS_CACHE: dict = {"ts": 0.0, "data": {}}
CUMDS_TTL = 900        # seconds between Influx queries (15 min)
CUMDS_MAX_TICKERS = 50  # cap query size regardless of --top — protects InfluxDB
CUMDS_ENABLED = True    # set False by --no-cumds

INFLUX_URL = os.environ.get("OPBOARD_INFLUX_URL", "http://localhost:8086")
INFLUX_TOKEN = os.environ.get("INFLUX_TOKEN", "element47-dev-token")
INFLUX_ORG = os.environ.get("INFLUX_ORG", "element47")
INFLUX_BUCKET = os.environ.get("INFLUX_BUCKET", "sma_engine")


def read_cumulative_ds(tickers) -> dict[str, float]:
    """
    Max 7-day cumulative deciseconds per ticker (its strongest level).
    Queries InfluxDB for the given tickers only — keeps the Flux fast.
    Pass tickers in display order; only the first CUMDS_MAX_TICKERS are queried
    to protect InfluxDB while the engines are writing.
    Cached for CUMDS_TTL seconds; fails silently to last-known data.
    """
    if not CUMDS_ENABLED:
        return {}
    now = time.time()
    if now - _CUMDS_CACHE["ts"] < CUMDS_TTL and _CUMDS_CACHE["data"]:
        return _CUMDS_CACHE["data"]
    if not tickers:
        return _CUMDS_CACHE["data"]
    tickers = list(dict.fromkeys(tickers))[:CUMDS_MAX_TICKERS]
    try:
        from influxdb_client import InfluxDBClient
    except ImportError:
        return {}
    try:
        client = InfluxDBClient(url=INFLUX_URL, token=INFLUX_TOKEN,
                                org=INFLUX_ORG, timeout=45_000)
        # OR chain (not contains) — allows Influx predicate pushdown, which makes
        # this orders of magnitude faster. Safe below ~100 tickers; the screen
        # shows at most --top rows so we're well under the nesting limit.
        or_chain = " or ".join(f'r.ticker == "{t}"' for t in sorted(tickers))
        flux = f"""
from(bucket: "{INFLUX_BUCKET}")
  |> range(start: -7d)
  |> filter(fn: (r) => r._measurement == "hits" and r._field == "deciseconds")
  |> filter(fn: (r) => {or_chain})
  |> group(columns: ["ticker", "timeframe", "outfit_id", "sma_period"])
  |> sum(column: "_value")
"""
        tables = client.query_api().query(flux, org=INFLUX_ORG)
        best: dict[str, float] = {}
        for tb in tables:
            for rec in tb.records:
                t = rec.values.get("ticker", "")
                v = float(rec.get_value() or 0)
                if v > best.get(t, 0):
                    best[t] = v
        client.close()
        _CUMDS_CACHE["ts"] = now
        _CUMDS_CACHE["data"] = best
        return best
    except Exception:
        # Back off 60s before retrying so a slow Influx doesn't stall every render
        _CUMDS_CACHE["ts"] = now - CUMDS_TTL + 60
        return _CUMDS_CACHE["data"]


def fmt_ds(v: float) -> str:
    if v >= 1e9: return f"{v/1e9:.1f}B"
    if v >= 1e6: return f"{v/1e6:.1f}M"
    if v >= 1e3: return f"{v/1e3:.0f}k"
    return f"{v:.0f}" if v else "—"


# ── Zero gamma per ticker (computed locally from option chains, cached daily) ─

GAMMA_ENABLED = False          # set by --gamma
GAMMA_MAX_TICKERS = 25         # only the top rows get gamma
GAMMA_CACHE_FILE = BASE / "output" / "operator" / ".opboard_gex_cache.json"

_GAMMA_DATA: dict[str, dict] = {}    # ticker -> gex result (thread-shared)
_GAMMA_STATUS: dict = {"running": False, "done": 0, "total": 0}


def _load_gamma_cache() -> None:
    """Load today's cached gamma results into _GAMMA_DATA."""
    import json
    from datetime import date
    if not GAMMA_CACHE_FILE.exists():
        return
    try:
        cache = json.loads(GAMMA_CACHE_FILE.read_text())
        today = date.today().isoformat()
        for t, r in cache.items():
            if r.get("cache_date") == today and not r.get("error"):
                _GAMMA_DATA[t] = r
    except Exception:
        pass


def _gamma_worker(ticker_prices: list[tuple[str, float]]) -> None:
    """Background thread: compute zero gamma for each ticker, update cache."""
    import json
    import sys as _sys
    from datetime import date
    _sys.path.insert(0, str(BASE / "market_overlay"))
    try:
        from index_gex import _compute_gex
    except ImportError:
        _GAMMA_STATUS["running"] = False
        return
    try:
        cache = (json.loads(GAMMA_CACHE_FILE.read_text())
                 if GAMMA_CACHE_FILE.exists() else {})
    except Exception:
        cache = {}
    today = date.today().isoformat()
    todo = [(t, px) for t, px in ticker_prices
            if not (cache.get(t, {}).get("cache_date") == today
                    and not cache[t].get("error"))]
    _GAMMA_STATUS.update(running=True, done=0, total=len(todo))
    for t, px in todo:
        res = _compute_gex(t, px)
        res["cache_date"] = today
        cache[t] = res
        if not res.get("error"):
            _GAMMA_DATA[t] = res
        _GAMMA_STATUS["done"] += 1
        try:
            GAMMA_CACHE_FILE.parent.mkdir(parents=True, exist_ok=True)
            GAMMA_CACHE_FILE.write_text(json.dumps(cache, default=str))
        except Exception:
            pass
    _GAMMA_STATUS["running"] = False


def start_gamma_thread(rows: list) -> None:
    """Kick off background gamma computation for the top rows (once)."""
    import threading
    if _GAMMA_STATUS["running"]:
        return
    targets = [(r.ticker, r.price) for r in rows[:GAMMA_MAX_TICKERS]
               if r.price and r.ticker not in _GAMMA_DATA]
    if not targets:
        return
    _GAMMA_STATUS["running"] = True
    threading.Thread(target=_gamma_worker, args=(targets,), daemon=True).start()


# ── Per-engine 7-day appearance history (from archived output files) ──────────

_H7_CACHE: dict = {"ts": 0.0, "data": {}, "totals": (0, 0, 0)}
_H7_FILE_MEMO: dict = {}   # path -> frozenset of tickers (files never change)
H7_TTL = 600               # rescan directories every 10 min


def _tickers_in_snapshot(path: Path) -> frozenset:
    key = str(path)
    if key not in _H7_FILE_MEMO:
        try:
            with open(path) as fp:
                _H7_FILE_MEMO[key] = frozenset(
                    row["ticker"] for row in csv.DictReader(fp))
        except Exception:
            _H7_FILE_MEMO[key] = frozenset()
    return _H7_FILE_MEMO[key]


def _tickers_in_xlsx(path: Path, ticker_col: int, state_col: int = -1,
                     states: tuple = ()) -> frozenset:
    """Tickers in an xlsx output file. If state_col >= 0, only rows whose
    state is in `states` count (used for V3 STRONG/HOLD)."""
    key = str(path)
    if key not in _H7_FILE_MEMO:
        found = set()
        try:
            wb = openpyxl.load_workbook(path, read_only=True)
            ws = wb["V3 Signals"] if "V3 Signals" in wb.sheetnames else wb.active
            for row in ws.iter_rows(values_only=True, min_row=2):
                if not row or len(row) <= ticker_col or not row[ticker_col]:
                    continue
                if state_col >= 0 and str(row[state_col] or "") not in states:
                    continue
                found.add(str(row[ticker_col]))
            wb.close()
        except Exception:
            pass
        _H7_FILE_MEMO[key] = frozenset(found)
    return _H7_FILE_MEMO[key]


def read_history_7d() -> tuple[dict, tuple]:
    """
    Per-engine appearance counts over the last 7 days of archived outputs.
    Returns ({ticker: [main_n, norm_n, v3_n]}, (main_total, norm_total, v3_total)).
    V3 counts only STRONG/HOLD appearances.
    """
    now = time.time()
    if now - _H7_CACHE["ts"] < H7_TTL and _H7_CACHE["data"]:
        return _H7_CACHE["data"], _H7_CACHE["totals"]

    cutoff = now - 7 * 86400
    counts: dict[str, list] = {}

    def recent(directory: Path, pattern: str) -> list[Path]:
        if not directory.exists():
            return []
        return [f for f in directory.glob(pattern) if f.stat().st_mtime > cutoff]

    snap_files = recent(BASE / "output" / "snapshots", "snapshot_*.csv")
    norm_files = recent(PATHS["normalized"], "normalized_*.xlsx")
    v3_files = recent(PATHS["v3"], "v3_*.xlsx")

    for f in snap_files:
        for t in _tickers_in_snapshot(f):
            counts.setdefault(t, [0, 0, 0])[0] += 1
    for f in norm_files:
        for t in _tickers_in_xlsx(f, ticker_col=1):
            counts.setdefault(t, [0, 0, 0])[1] += 1
    for f in v3_files:
        for t in _tickers_in_xlsx(f, ticker_col=2, state_col=1,
                                  states=("STRONG", "HOLD")):
            counts.setdefault(t, [0, 0, 0])[2] += 1

    totals = (len(snap_files), len(norm_files), len(v3_files))
    _H7_CACHE.update(ts=now, data=counts, totals=totals)
    return counts, totals


# ── Combine (no scoring — raw merge) ──────────────────────────────────────────

def combine(main, norm, v3, disc, conf, trades, bt, ohlc=None) -> list[OperatorRow]:
    ohlc = ohlc or {}
    all_tickers = (set(main) | set(norm) | set(v3) | set(disc)
                   | set(conf) | set(trades) | set(bt))
    rows: list[OperatorRow] = []

    for t in all_tickers:
        r = OperatorRow(ticker=t)
        n_src = 0

        if t in main:
            m = main[t]
            r.main_rank, r.main_tf = m["rank"], m["tf"]
            r.main_outfit, r.main_hits, r.main_conv = m["outfit"], m["hits"], m["conv"]
            n_src += 1
        if t in norm:
            n = norm[t]
            r.norm_rank, r.norm_tf, r.norm_conv = n["rank"], n["tf"], n["conv"]
            r.norm_entry = n["entry"]
            n_src += 1
        if t in v3:
            v = v3[t]
            r.v3_rank, r.v3_state, r.v3_tf = v["rank"], v["state"], v["tf"]
            r.v3_grade, r.v3_xtf = v["grade"], v["x_tf"]
            n_src += 1
        if t in disc:
            d = disc[t]
            r.disc_tf = d.get("timeframe", "")
            r.disc_sma = f"MA{d.get('sma_period', '')}"
            r.disc_dir = d.get("direction", "")
            try:
                r.disc_sma_value = float(d.get("sma_value", 0) or 0)
                r.disc_close = float(d.get("close", 0) or 0)
                r.disc_high = float(d.get("high", 0) or 0)
                r.disc_low = float(d.get("low", 0) or 0)
            except (ValueError, TypeError):
                pass
            n_src += 1
        if t in conf:
            r.conf_score = int(conf[t].get("score", 0) or 0)
            n_src += 1
        if t in trades:
            tr = trades[t]
            r.trade_side = tr.get("side", "")
            r.trade_conf = tr.get("confidence", "")
            try:
                r.trade_entry = float(tr.get("entry", 0) or 0)
                r.trade_stop = float(tr.get("stop", 0) or 0)
                r.trade_rr = float(tr.get("rr", 0) or 0)
            except (ValueError, TypeError):
                pass
            n_src += 1
        if t in bt:
            b = bt[t]
            try:
                r.bt_sharpe = float(b.get("sharpe", 0) or 0)
                r.bt_winrate = float(b.get("win_rate", 0) or 0)
            except (ValueError, TypeError):
                pass
            n_src += 1
        if t in ohlc:
            o = ohlc[t]
            try:
                r.ohlc_o = float(o.get("open", 0) or 0)
                r.ohlc_h = float(o.get("high", 0) or 0)
                r.ohlc_l = float(o.get("low", 0) or 0)
                r.ohlc_c = float(o.get("close", 0) or 0)
                r.ohlc_tf = str(o.get("timeframe", "") or "")
            except (ValueError, TypeError):
                pass

        r.sources = n_src
        rows.append(r)

    # Sort: most sources first, then best main rank (0 = not present, sort last)
    rows.sort(key=lambda r: (-r.sources, r.main_rank if r.main_rank else 10**9))
    return rows


# ── File freshness ────────────────────────────────────────────────────────────

def file_ages() -> dict[str, str]:
    checks = {
        "main":  (PATHS["main"], None),
        "norm":  (PATHS["normalized"], "normalized_*.xlsx"),
        "v3":    (PATHS["v3"], "v3_*.xlsx"),
        "disc":  (PATHS["discovery"], "discovery_*.csv"),
        "conf":  (PATHS["confluence"], "confluence_*.csv"),
        "trade": (PATHS["trades"], "trades_*.csv"),
        "bt":    (PATHS["backtest"], "backtest_*.csv"),
    }
    now = time.time()
    out: dict[str, str] = {}
    for name, (path, pattern) in checks.items():
        f = latest_file(path, pattern) if pattern else (path if path.exists() else None)
        if not f:
            out[name] = "[dim]—[/dim]"
            continue
        age_min = (now - f.stat().st_mtime) / 60
        if age_min < 60:
            out[name] = f"[green]{age_min:.0f}m[/green]"
        elif age_min < 240:
            out[name] = f"[yellow]{age_min/60:.1f}h[/yellow]"
        else:
            out[name] = f"[red]{age_min/60:.0f}h[/red]"
    return out


# ── UI ────────────────────────────────────────────────────────────────────────

STATE_STYLE = {"STRONG": "bright_green", "HOLD": "yellow", "WEAK": "red", "IGNORE": "dim"}
CONV_STYLE = {"4/4": "bright_green", "3/4": "green", "2/4": "yellow", "1/4": "dim"}


def build_header(ages: dict[str, str]) -> Panel:
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    age_str = "  ".join(f"{k} {v}" for k, v in ages.items())
    return Panel(
        Text(f"OPERATOR — RAW CROSS-ENGINE BOARD  ✦  {now}", style="bold white",
             justify="center"),
        subtitle=age_str,
        border_style="bright_white",
        padding=(0, 1),
    )


def build_table(rows: list[OperatorRow], top_n: int, min_src: int,
                only_ticker: Optional[str], sort: str = "src") -> Panel:
    table = Table(
        box=box.SIMPLE_HEAD, padding=(0, 1), show_header=True,
        header_style="bold white", expand=True,
    )
    table.add_column("Ticker", width=7, style="bold")
    table.add_column("Src", width=4)
    table.add_column("Main", width=14)          # rank·tf·conv·hits
    table.add_column("Outfit", width=17, style="dim")
    table.add_column("Norm", width=10)          # rank·conv
    table.add_column("V3", width=13)            # state·grade·xtf
    table.add_column("Disc", width=13)          # tf MA dir
    table.add_column("Cf", width=3)
    table.add_column("Sharpe/WR", width=10)
    table.add_column("Trade", width=17)         # side entry→stop RR
    table.add_column("7d M·N·V", width=9)       # cycles seen per engine, 7 days
    table.add_column("CumDS", width=6, justify="right")
    table.add_column("L/S", width=3)            # price above (L) or below (S) entry
    table.add_column("0γ", width=7)             # above/below zero gamma + dist%
    table.add_column("OHLC (tf candle)", width=19, justify="right")

    shown = [r for r in rows if r.sources >= min_src]
    if only_ticker:
        shown = [r for r in shown if r.ticker.upper() == only_ticker.upper()]
    shown = shown[:top_n]

    # Cumulative deciseconds — top rows only, in display order (protects Influx)
    cum = read_cumulative_ds([r.ticker for r in shown])
    # 7-day per-engine appearance history (cached, memoized per file)
    h7, h7_totals = read_history_7d()
    for r in shown:
        r.cum_ds = cum.get(r.ticker, 0.0)
        m, n, v = h7.get(r.ticker, [0, 0, 0])
        r.h7_main, r.h7_norm, r.h7_v3 = m, n, v

    if sort == "cumds":
        shown.sort(key=lambda r: -r.cum_ds)

    # Zero gamma — kick off background computation for the rows on screen
    if GAMMA_ENABLED:
        start_gamma_thread(shown)

    for r in shown:
        src_style = ("bright_green" if r.sources >= 5 else
                     "green" if r.sources >= 3 else
                     "yellow" if r.sources == 2 else "dim")

        # Main
        if r.main_rank:
            main_str = f"#{r.main_rank} {r.main_tf} {r.main_conv} {r.main_hits}"
            main_style = CONV_STYLE.get(r.main_conv, "white")
        else:
            main_str, main_style = "—", "dim"

        outfit_str = r.main_outfit if r.main_outfit else "—"

        # Norm
        if r.norm_rank:
            norm_str = f"#{r.norm_rank} {r.norm_conv}"
            norm_style = CONV_STYLE.get(r.norm_conv, "white")
        else:
            norm_str, norm_style = "—", "dim"

        # V3
        if r.v3_state:
            v3_str = f"{r.v3_state} {r.v3_grade}"
            if r.v3_xtf:
                v3_str += f" x{r.v3_xtf}"
            v3_style = STATE_STYLE.get(r.v3_state, "white")
        else:
            v3_str, v3_style = "—", "dim"

        # Discovery
        if r.disc_tf:
            arrow = "↑" if r.disc_dir == "from_below" else ("↓" if r.disc_dir == "from_above" else "")
            disc_str = f"{r.disc_tf} {r.disc_sma}{arrow}"
            disc_style = ("bright_green" if r.disc_tf in ("1mo", "1w", "1d")
                          else "green" if r.disc_tf in ("4h", "2h") else "yellow")
        else:
            disc_str, disc_style = "—", "dim"

        # Confluence
        conf_str = str(r.conf_score) if r.conf_score else "—"
        conf_style = ("bright_green" if r.conf_score >= 3 else
                      "green" if r.conf_score == 2 else "dim")

        # Backtest
        if r.bt_sharpe or r.bt_winrate:
            bt_str = f"{r.bt_sharpe:.1f}/{r.bt_winrate*100:.0f}%"
            bt_style = ("bright_green" if r.bt_sharpe > 15 else
                        "green" if r.bt_sharpe > 8 else "white")
        else:
            bt_str, bt_style = "—", "dim"

        # Trade
        if r.trade_side:
            arrow = "↑" if r.trade_side == "BUY" else "↓"
            trade_str = f"{arrow}{r.trade_conf[:3]} {r.trade_entry:.2f}→{r.trade_stop:.2f} R{r.trade_rr:.1f}"
            trade_style = ("bright_green" if r.trade_conf == "HIGH" else
                           "yellow" if r.trade_conf == "MEDIUM" else "dim")
        else:
            trade_str, trade_style = "—", "dim"

        # 7-day per-engine history
        if r.h7_main or r.h7_norm or r.h7_v3:
            h7_str = f"{r.h7_main}·{r.h7_norm}·{r.h7_v3}"
            persistent = sum(1 for x in (r.h7_main, r.h7_norm, r.h7_v3) if x >= 5)
            h7_style = ("bright_green" if persistent >= 2 else
                        "green" if persistent == 1 else "white")
        else:
            h7_str, h7_style = "—", "dim"

        # Cumulative deciseconds
        cum_str = fmt_ds(r.cum_ds)
        cum_style = ("bright_green" if r.cum_ds >= 1e9 else
                     "green" if r.cum_ds >= 1e6 else
                     "yellow" if r.cum_ds >= 1e3 else "dim")

        # L/S — price above or below the entry reference
        ls = r.long_short
        if ls == "L":
            ls_str, ls_style = "L", "bright_green"
        elif ls == "S":
            ls_str, ls_style = "S", "red"
        else:
            ls_str, ls_style = "—", "dim"

        # Zero gamma — above (A) / below (B) + distance %
        g = _GAMMA_DATA.get(r.ticker)
        if g and g.get("zero_gamma"):
            above = r.price > g["zero_gamma"] if r.price else \
                    g.get("regime", "").startswith("ABOVE")
            dist = ((r.price - g["zero_gamma"]) / g["zero_gamma"] * 100
                    if r.price else g.get("dist_pct", 0))
            gam_str = f"{'A' if above else 'B'}{dist:+.1f}"
            gam_style = "bright_green" if above else "red"
        elif GAMMA_ENABLED and _GAMMA_STATUS["running"]:
            gam_str, gam_style = "…", "dim"
        else:
            gam_str, gam_style = "—", "dim"

        # OHLC (two lines: O H / L C + the candle's timeframe)
        if r.ohlc_c:
            tf_tag = f" {r.ohlc_tf}" if r.ohlc_tf else ""
            ohlc_str = (f"{r.ohlc_o:.2f} {r.ohlc_h:.2f}\n"
                        f"{r.ohlc_l:.2f} {r.ohlc_c:.2f}{tf_tag}")
        elif r.price:
            ohlc_str = f"{r.price:.2f}"
        else:
            ohlc_str = "—"

        table.add_row(
            r.ticker,
            Text(f"{r.sources}/7", style=src_style),
            Text(main_str, style=main_style),
            outfit_str,
            Text(norm_str, style=norm_style),
            Text(v3_str, style=v3_style),
            Text(disc_str, style=disc_style),
            Text(conf_str, style=conf_style),
            Text(bt_str, style=bt_style),
            Text(trade_str, style=trade_style),
            Text(h7_str, style=h7_style),
            Text(cum_str, style=cum_style),
            Text(ls_str, style=ls_style),
            Text(gam_str, style=gam_style),
            ohlc_str,
        )

    total = len(rows)
    mt, nt, vt = h7_totals
    return Panel(
        table,
        title="[bold cyan]MASTER SIGNAL BOARD[/bold cyan]",
        subtitle=f"[dim]{len(shown)} shown · {total} tickers across all sources"
                 f" · min-src {min_src}"
                 f" · 7d cycles: main {mt} norm {nt} v3 {vt}[/dim]",
        border_style="cyan",
    )


# ── Export (spreadsheet + historical log — used by EOD pipeline) ──────────────

EXPORT_HEADER = [
    "timestamp_utc", "ticker", "sources",
    "main_rank", "main_tf", "main_conv", "main_hits", "outfit",
    "norm_rank", "norm_conv",
    "v3_rank", "v3_state", "v3_grade", "v3_xtf",
    "disc_tf", "disc_sma", "disc_dir",
    "conf_score", "bt_sharpe", "bt_winrate",
    "trade_side", "trade_conf", "trade_entry", "trade_stop", "trade_rr",
    "h7_main", "h7_norm", "h7_v3", "cum_ds",
    "entry_ref", "long_short",
    "zero_gamma", "gamma_ab", "gamma_dist_pct",
    "open", "high", "low", "close",
]


def export(min_src: int = 1) -> None:
    """
    One-shot export: write the full combined board to
      output/operator/operator_<ts>.xlsx   (snapshot, accumulates historically)
      output/operator/operator_log.csv     (flat append-only log)
    then exit. No UI, no live loop. Used by eod.sh / fullrun.sh before sheets sync.
    """
    rows = combine(read_main(), read_normalized(), read_v3(), read_discovery(),
                   read_confluence(), read_trades(), read_backtest(), read_ohlc())
    shown = [r for r in rows if r.sources >= min_src]

    cum = read_cumulative_ds([r.ticker for r in shown])
    h7, _ = read_history_7d()
    for r in shown:
        r.cum_ds = cum.get(r.ticker, 0.0)
        m, n, v = h7.get(r.ticker, [0, 0, 0])
        r.h7_main, r.h7_norm, r.h7_v3 = m, n, v

    ts = datetime.now(timezone.utc)
    ts_str = ts.strftime("%Y-%m-%d_%H-%M-%S")
    out_dir = BASE / "output" / "operator"
    out_dir.mkdir(parents=True, exist_ok=True)

    def row_values(r: OperatorRow) -> list:
        g = _GAMMA_DATA.get(r.ticker, {})
        zg = g.get("zero_gamma", "")
        if zg and r.price:
            gamma_ab = "A" if r.price > zg else "B"
            gamma_dist = round((r.price - zg) / zg * 100, 2)
        else:
            gamma_ab, gamma_dist = "", ""
        return [
            ts.isoformat(), r.ticker, r.sources,
            r.main_rank, r.main_tf, r.main_conv, r.main_hits, r.main_outfit,
            r.norm_rank, r.norm_conv,
            r.v3_rank, r.v3_state, r.v3_grade, r.v3_xtf,
            r.disc_tf, r.disc_sma, r.disc_dir,
            r.conf_score, r.bt_sharpe, r.bt_winrate,
            r.trade_side, r.trade_conf, r.trade_entry, r.trade_stop, r.trade_rr,
            r.h7_main, r.h7_norm, r.h7_v3, r.cum_ds,
            r.entry_ref, r.long_short,
            zg, gamma_ab, gamma_dist,
            r.ohlc_o, r.ohlc_h, r.ohlc_l, r.ohlc_c,
        ]

    # ── xlsx snapshot ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Operator"
    ws.append(EXPORT_HEADER)
    for r in shown:
        ws.append(row_values(r))
    xlsx_path = out_dir / f"operator_{ts_str}.xlsx"
    wb.save(xlsx_path)

    # ── CSV log append ───────────────────────────────────────────────────────
    log_path = out_dir / "operator_log.csv"
    new_file = not log_path.exists()
    with open(log_path, "a", newline="") as fp:
        w = csv.writer(fp)
        if new_file:
            w.writerow(EXPORT_HEADER)
        for r in shown:
            w.writerow(row_values(r))

    print(f"operator export: {len(shown)} tickers → {xlsx_path.name} "
          f"(+{len(shown)} rows to operator_log.csv)")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Operator dashboard — raw cross-engine board")
    parser.add_argument("--top", type=int, default=25, help="Rows to display (default 25)")
    parser.add_argument("--refresh", type=int, default=30, help="Refresh seconds (default 30)")
    parser.add_argument("--min-src", type=int, default=1,
                        help="Only show tickers in at least N sources (default 1)")
    parser.add_argument("--ticker", type=str, default=None, help="Show only this ticker")
    parser.add_argument("--sort", type=str, default="cumds", choices=["src", "cumds"],
                        help="cumds = most cumulative deciseconds first (default), src = most sources first")
    parser.add_argument("--no-cumds", action="store_true",
                        help="Disable InfluxDB cumulative deciseconds queries entirely "
                             "(use while engines are mid-scan to keep Influx load at zero)")
    parser.add_argument("--export", action="store_true",
                        help="One-shot: write board to output/operator/ (xlsx + csv log) and exit")
    parser.add_argument("--gamma", action="store_true",
                        help="Compute zero gamma per ticker (top rows) from option chains. "
                             "First run of the day takes a few minutes in the background; "
                             "cached for the trading day after that.")
    args = parser.parse_args()

    if args.no_cumds:
        global CUMDS_ENABLED
        CUMDS_ENABLED = False

    global GAMMA_ENABLED
    GAMMA_ENABLED = args.gamma
    _load_gamma_cache()   # cached results show even without --gamma

    if args.export:
        export(min_src=args.min_src)
        return

    console = Console()
    layout = Layout()
    layout.split_column(
        Layout(name="header", size=3),
        Layout(name="board"),
    )

    try:
        with Live(layout, console=console, refresh_per_second=2, screen=True):
            while True:
                main_d = read_main()
                norm_d = read_normalized()
                v3_d = read_v3()
                disc_d = read_discovery()
                conf_d = read_confluence()
                tr_d = read_trades()
                bt_d = read_backtest()
                ohlc_d = read_ohlc()

                rows = combine(main_d, norm_d, v3_d, disc_d, conf_d, tr_d, bt_d, ohlc_d)

                layout["header"].update(build_header(file_ages()))
                layout["board"].update(
                    build_table(rows, args.top, args.min_src, args.ticker, args.sort))

                time.sleep(args.refresh)
    except KeyboardInterrupt:
        pass
    finally:
        console.print("\n[dim]Operator dashboard stopped.[/dim]")


if __name__ == "__main__":
    main()
