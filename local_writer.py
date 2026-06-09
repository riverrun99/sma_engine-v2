"""
local_writer.py — local file output, no cloud required.

Always-on fallback for subscribers who don't set up Google Sheets. Produces
two files in the configured output directory (default: ./output):

  signals_current.xlsx
    Overwritten every cycle. Contains a single sheet with the snapshot
    layout: top signal block, ranked leaderboard, system states block,
    regime cell. Also includes a Performance tab tracking price moves
    since each ticker's first appearance in the ranked output.
    Mirrors what the Google Sheets "Current" tab shows.

  signals_log.csv
    Append-only. One row per scan cycle. New rows are added to the bottom;
    existing rows are never modified. Mirrors what the Google Sheets "Log"
    tab shows. CSV format chosen because it's universal — opens in Excel,
    Numbers, Google Sheets, any text editor.

  ranked_log.csv
    Append-only. One row per ranked entry per cycle.

  price_tracker.json
    Persistent store of each ticker's first-seen price and timestamp.
    Used to compute % change since initial scan appearance.

If openpyxl is unavailable, the XLSX file is skipped and only the CSV is
written. The engine never crashes on output failures.
"""

from __future__ import annotations

import os
import csv
import json
import logging
from pathlib import Path
from datetime import datetime, timezone
from typing import Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Schema for the append-only ranked log. One row per ranked entry per cycle.
RANKED_LOG_COLUMNS = [
    "timestamp_utc",
    "rank", "ticker", "timeframe", "outfit_periods",
    "hit_count", "convergence", "rank_score",
]

# Schema for the append-only log. Order matters — first row of the CSV.
LOG_COLUMNS = [
    "timestamp_utc",
    "top_ticker", "top_timeframe", "top_outfit_id", "top_outfit_periods",
    "top_entry_price", "top_offset", "top_hit_count", "top_convergence",
    "conv_ohlc", "conv_close", "conv_parm", "conv_timeseries",
    "sp500", "nasdaq", "dow", "vix", "svix", "russell2000", "russell3000", "semis",
    "regime_label",
]


class LocalWriter:
    """Writes Current snapshot (xlsx) + append-only Log (csv) to local disk."""

    def __init__(self, output_dir: str = "./output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.log_path = self.output_dir / "signals_log.csv"
        self.ranked_log_path = self.output_dir / "ranked_log.csv"
        self.current_xlsx_path = self.output_dir / "signals_current.xlsx"
        self.price_tracker_path = self.output_dir / "price_tracker.json"
        self.snapshots_dir = self.output_dir / "snapshots"
        self.snapshots_dir.mkdir(parents=True, exist_ok=True)
        self.ohlc_log_path = self.output_dir / "ohlc_log.csv"
        self._price_tracker: dict = self._load_price_tracker()
        self._ensure_log_header()
        self._ensure_ranked_log_header()
        self._ensure_ohlc_log_header()

    # ── Price tracker ─────────────────────────────────────────────────────────

    def _load_price_tracker(self) -> dict:
        """Load persisted price tracker from disk."""
        if self.price_tracker_path.exists():
            try:
                with open(self.price_tracker_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                logging.warning(f"LocalWriter: could not load price tracker: {e}")
        return {}

    def _save_price_tracker(self) -> None:
        """Persist price tracker to disk."""
        try:
            with open(self.price_tracker_path, "w", encoding="utf-8") as f:
                json.dump(self._price_tracker, f, indent=2)
        except Exception as e:
            logging.warning(f"LocalWriter: could not save price tracker: {e}")

    def update_price_tracker(self, top_n: list, ts: datetime) -> None:
        """
        For each ticker in top_n, record first-seen price and timestamp.
        On subsequent appearances, update current price only.
        Uses entry_price (the SMA level at hit time) as the price proxy.
        """
        ts_str = ts.isoformat()
        changed = False
        # Build a deduplicated map: ticker -> best entry_price (highest rank = rank 1)
        seen: dict[str, float] = {}
        for entry in top_n:
            ticker = entry.get("ticker", "")
            price = entry.get("entry_price")
            if not ticker or price is None:
                continue
            # Take price from the highest-ranked (lowest rank number) appearance
            if ticker not in seen:
                seen[ticker] = float(price)

        for ticker, price in seen.items():
            if ticker not in self._price_tracker:
                # First time we've seen this ticker — record it
                self._price_tracker[ticker] = {
                    "first_seen": ts_str,
                    "first_price": price,
                    "current_price": price,
                    "last_updated": ts_str,
                }
                changed = True
            else:
                # Update current price
                self._price_tracker[ticker]["current_price"] = price
                self._price_tracker[ticker]["last_updated"] = ts_str
                changed = True

        if changed:
            self._save_price_tracker()

    # ── Log headers ───────────────────────────────────────────────────────────

    def _ensure_log_header(self) -> None:
        """Create the log CSV with header row if it doesn't exist."""
        if not self.log_path.exists():
            try:
                with open(self.log_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(LOG_COLUMNS)
            except Exception as e:
                logging.warning(f"LocalWriter: could not create log file: {e}")

    def _ensure_ranked_log_header(self) -> None:
        """Create the ranked log CSV with header row if it doesn't exist."""
        if not self.ranked_log_path.exists():
            try:
                with open(self.ranked_log_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow(RANKED_LOG_COLUMNS)
            except Exception as e:
                logging.warning(f"LocalWriter: could not create ranked log file: {e}")

    def _ensure_ohlc_log_header(self) -> None:
        """Create the OHLC log CSV with header row if it doesn't exist."""
        if not self.ohlc_log_path.exists():
            try:
                with open(self.ohlc_log_path, "w", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerow([
                        "timestamp_utc", "ticker", "timeframe", "outfit",
                        "open", "high", "low", "close", "volume",
                        "hit_count", "convergence", "rank_score", "first_seen"
                    ])
            except Exception as e:
                logging.warning(f"LocalWriter: could not create ohlc log: {e}")

    def append_ohlc_log(self, top_n: list, candle_cache: dict, ts: datetime) -> None:
        """
        For each unique ticker in top_n, look up the latest candle bar
        and record OHLC + volume. Only logs tickers appearing for the first time
        (first_seen=1) or every appearance (first_seen=0).
        """
        ts_str = ts.isoformat()
        try:
            seen: set = set()
            rows = []
            for entry in top_n:
                ticker = entry.get("ticker", "")
                tf = entry.get("timeframe", "")
                if not ticker or ticker in seen:
                    continue
                seen.add(ticker)

                df = candle_cache.get((ticker, tf))
                if df is None or len(df) == 0:
                    continue

                latest = df.iloc[-1]
                periods_str = "/".join(str(p) for p in entry.get("outfit_periods", []))
                first_seen = 1 if ticker not in self._price_tracker else 0

                rows.append([
                    ts_str,
                    ticker,
                    tf,
                    periods_str,
                    round(float(latest.get("open", 0) or 0), 4),
                    round(float(latest.get("high", 0) or 0), 4),
                    round(float(latest.get("low", 0) or 0), 4),
                    round(float(latest.get("close", 0) or 0), 4),
                    int(latest.get("volume", 0) or 0),
                    entry.get("hit_count", ""),
                    entry.get("convergence", ""),
                    entry.get("rank_score", ""),
                    first_seen,
                ])

            if rows:
                with open(self.ohlc_log_path, "a", newline="", encoding="utf-8") as f:
                    writer = csv.writer(f)
                    writer.writerows(rows)
        except Exception as e:
            logging.warning(f"LocalWriter: ohlc log append failed: {e}")

    def write_snapshot(self, top_n: list, ts: datetime) -> None:
        """
        Write a timestamped snapshot file to output/snapshots/.
        One file per cycle. Lists each unique ticker once (best rank),
        with timeframe and SMA outfit. Simple and human-readable.
        """
        ts_str = ts.strftime("%Y-%m-%d_%H-%M-%S")
        filename = self.snapshots_dir / f"snapshot_{ts_str}.csv"
        try:
            seen: dict[str, dict] = {}
            for entry in top_n:
                ticker = entry.get("ticker", "")
                if not ticker or ticker in seen:
                    continue
                seen[ticker] = entry

            with open(filename, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["timestamp_utc", "ticker", "timeframe", "outfit", "hits", "convergence", "score"])
                for ticker, entry in seen.items():
                    periods_str = "/".join(str(p) for p in entry.get("outfit_periods", []))
                    writer.writerow([
                        ts.isoformat(),
                        ticker,
                        entry.get("timeframe", ""),
                        periods_str,
                        entry.get("hit_count", ""),
                        entry.get("convergence", ""),
                        entry.get("rank_score", ""),
                    ])
        except Exception as e:
            logging.warning(f"LocalWriter: snapshot write failed: {e}")

    def append_ranked_log(
        self,
        top_n: list,
        ts: Optional[datetime] = None,
    ) -> None:
        """Append one row per ranked entry to ranked_log.csv."""
        ts = ts or datetime.now(timezone.utc)
        ts_str = ts.isoformat()
        try:
            with open(self.ranked_log_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                for entry in top_n:
                    periods_str = "/".join(str(p) for p in entry.get("outfit_periods", []))
                    writer.writerow([
                        ts_str,
                        entry.get("rank", ""),
                        entry.get("ticker", ""),
                        entry.get("timeframe", ""),
                        periods_str,
                        entry.get("hit_count", ""),
                        entry.get("convergence", ""),
                        entry.get("rank_score", ""),
                    ])
        except Exception as e:
            logging.warning(f"LocalWriter: ranked log append failed: {e}")

    def _row_from_signal(
        self, signal: Optional[dict], systems: list, regime_label: Optional[str],
        ts: datetime,
    ) -> list:
        """Flatten signal + systems + regime into one row matching LOG_COLUMNS."""
        sys_state = {s.name: s.state for s in systems} if systems else {}

        if signal:
            conv = signal.get("convergence", {})
            periods_str = "/".join(str(p) for p in signal.get("outfit_periods", []))
            return [
                ts.isoformat(),
                signal.get("ticker", ""),
                signal.get("timeframe", ""),
                signal.get("outfit_id", ""),
                periods_str,
                signal.get("entry_price", ""),
                signal.get("offset_applied", ""),
                signal.get("hit_count", ""),
                conv.get("score", ""),
                int(bool(conv.get("ohlc_detection", False))),
                int(bool(conv.get("candle_close", False))),
                int(bool(conv.get("parm_price", False))),
                int(bool(conv.get("time_series", False))),
                sys_state.get("S&P 500", ""),
                sys_state.get("NASDAQ", ""),
                sys_state.get("Dow Jones", ""),
                sys_state.get("VIX", ""),
                sys_state.get("SVIX", ""),
                sys_state.get("Russell 2000", ""),
                sys_state.get("Russell 3000", ""),
                sys_state.get("Semiconductors", ""),
                regime_label or "",
            ]
        else:
            return [
                ts.isoformat(),
                "", "", "", "", "", "", "", "",
                "", "", "", "",
                sys_state.get("S&P 500", ""),
                sys_state.get("NASDAQ", ""),
                sys_state.get("Dow Jones", ""),
                sys_state.get("VIX", ""),
                sys_state.get("SVIX", ""),
                sys_state.get("Russell 2000", ""),
                sys_state.get("Russell 3000", ""),
                sys_state.get("Semiconductors", ""),
                regime_label or "",
            ]

    def append_log_row(
        self,
        signal: Optional[dict],
        systems: list,
        regime_label: Optional[str] = None,
        ts: Optional[datetime] = None,
    ) -> None:
        """Append one row to the log CSV."""
        ts = ts or datetime.now(timezone.utc)
        row = self._row_from_signal(signal, systems, regime_label, ts)
        try:
            with open(self.log_path, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(row)
        except Exception as e:
            logging.warning(f"LocalWriter: append failed: {e}")

    # ── XLSX ──────────────────────────────────────────────────────────────────

    def write_current_xlsx(
        self,
        signal: Optional[dict],
        top_n: list,
        systems: list,
        regime_label: Optional[str] = None,
        ts: Optional[datetime] = None,
    ) -> None:
        """Overwrite the Current xlsx with a snapshot of latest state."""
        if not OPENPYXL_AVAILABLE:
            return
        ts = ts or datetime.now(timezone.utc)
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Current"

            bold = Font(bold=True)
            header_fill = PatternFill(start_color="1F2937", end_color="1F2937",
                                       fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            green_font = Font(bold=True, color="00AA00")
            red_font = Font(bold=True, color="CC0000")

            row = 1

            # Header
            ws.cell(row=row, column=1, value="ELEMENT 47 — SMA ENGINE LIVE").font = Font(bold=True, size=14)
            row += 1
            ws.cell(row=row, column=1, value="Last update:")
            ws.cell(row=row, column=2, value=ts.isoformat())
            row += 2

            # Top signal block
            ws.cell(row=row, column=1, value="TOP SIGNAL").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            row += 1
            if signal:
                periods_str = "/".join(str(p) for p in signal.get("outfit_periods", []))
                pairs = [
                    ("Ticker", signal.get("ticker", "")),
                    ("Timeframe", signal.get("timeframe", "")),
                    ("Outfit", f"{periods_str} ({signal.get('outfit_name', '')})"),
                    ("Entry price", signal.get("entry_price", "")),
                    ("Offset", signal.get("offset_applied", "")),
                    ("Hit count", signal.get("hit_count", "")),
                    ("Convergence", signal.get("convergence", {}).get("score", "")),
                    ("Risk", signal.get("risk", "")),
                ]
                for label, value in pairs:
                    ws.cell(row=row, column=1, value=label).font = bold
                    ws.cell(row=row, column=2, value=value)
                    row += 1
            else:
                ws.cell(row=row, column=1, value="(no signal detected this cycle)")
                row += 1
            row += 1

            # Ranked leaderboard block
            ws.cell(row=row, column=1, value=f"TOP {len(top_n)} RANKED").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            row += 1
            headers = ["Rank", "Ticker", "TF", "Outfit", "Hits", "Conv", "Score"]
            for col_idx, h in enumerate(headers, start=1):
                ws.cell(row=row, column=col_idx, value=h).font = bold
            row += 1
            for entry in top_n:
                periods_str = "/".join(str(p) for p in entry.get("outfit_periods", []))
                ws.cell(row=row, column=1, value=entry.get("rank"))
                ws.cell(row=row, column=2, value=entry.get("ticker"))
                ws.cell(row=row, column=3, value=entry.get("timeframe"))
                ws.cell(row=row, column=4, value=periods_str)
                ws.cell(row=row, column=5, value=entry.get("hit_count"))
                ws.cell(row=row, column=6, value=entry.get("convergence"))
                ws.cell(row=row, column=7, value=entry.get("rank_score"))
                row += 1
            row += 1

            # System states block
            ws.cell(row=row, column=1, value="SYSTEM STATES").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            row += 1
            ws.cell(row=row, column=1, value="System").font = bold
            ws.cell(row=row, column=2, value="State").font = bold
            ws.cell(row=row, column=3, value="Note").font = bold
            row += 1
            for s in systems:
                ws.cell(row=row, column=1, value=s.name)
                ws.cell(row=row, column=2, value=s.state.upper())
                ws.cell(row=row, column=3, value=s.note)
                row += 1
            row += 1

            # Regime
            ws.cell(row=row, column=1, value="REGIME").font = header_font
            ws.cell(row=row, column=1).fill = header_fill
            ws.cell(row=row, column=2, value=regime_label or "(not yet computed)").font = bold

            # Column widths
            ws.column_dimensions["A"].width = 18
            ws.column_dimensions["B"].width = 28
            ws.column_dimensions["C"].width = 18
            ws.column_dimensions["D"].width = 30

            # ── Performance tab ───────────────────────────────────────────────
            if self._price_tracker:
                wp = wb.create_sheet(title="Performance")

                wp.cell(row=1, column=1, value="ELEMENT 47 — PRICE PERFORMANCE").font = Font(bold=True, size=14)
                wp.cell(row=2, column=1, value="Tracks price change since each ticker's first appearance in ranked output.")
                wp.cell(row=3, column=1, value="Last update:")
                wp.cell(row=3, column=2, value=ts.isoformat())

                perf_headers = ["Ticker", "First Seen", "First Price", "Current Price", "Change %", "Days Tracked"]
                for col_idx, h in enumerate(perf_headers, start=1):
                    c = wp.cell(row=5, column=col_idx, value=h)
                    c.font = header_font
                    c.fill = header_fill

                # Sort by % change descending
                perf_rows = []
                for ticker, data in self._price_tracker.items():
                    first_price = data.get("first_price", 0)
                    current_price = data.get("current_price", 0)
                    first_seen = data.get("first_seen", "")
                    if first_price and first_price > 0:
                        pct_change = ((current_price - first_price) / first_price) * 100
                    else:
                        pct_change = 0.0
                    try:
                        first_dt = datetime.fromisoformat(first_seen)
                        days = (ts - first_dt).days
                    except Exception:
                        days = 0
                    perf_rows.append((ticker, first_seen, first_price, current_price, pct_change, days))

                perf_rows.sort(key=lambda x: x[4], reverse=True)

                for r_idx, (ticker, first_seen, first_price, current_price, pct_change, days) in enumerate(perf_rows, start=6):
                    wp.cell(row=r_idx, column=1, value=ticker).font = bold
                    # Format first_seen to date only
                    try:
                        first_seen_str = datetime.fromisoformat(first_seen).strftime("%Y-%m-%d %H:%M")
                    except Exception:
                        first_seen_str = first_seen
                    wp.cell(row=r_idx, column=2, value=first_seen_str)
                    wp.cell(row=r_idx, column=3, value=round(first_price, 4))
                    wp.cell(row=r_idx, column=4, value=round(current_price, 4))
                    pct_cell = wp.cell(row=r_idx, column=5, value=round(pct_change, 2))
                    pct_cell.font = green_font if pct_change >= 0 else red_font
                    wp.cell(row=r_idx, column=6, value=days)

                wp.column_dimensions["A"].width = 10
                wp.column_dimensions["B"].width = 20
                wp.column_dimensions["C"].width = 14
                wp.column_dimensions["D"].width = 14
                wp.column_dimensions["E"].width = 12
                wp.column_dimensions["F"].width = 14

            wb.save(self.current_xlsx_path)
            # Also save a timestamped archive copy
            archive_dir = self.output_dir / "xlsx_archive"
            archive_dir.mkdir(parents=True, exist_ok=True)
            ts_str = ts.strftime("%Y-%m-%d_%H-%M-%S")
            archive_path = archive_dir / f"signals_{ts_str}.xlsx"
            wb.save(archive_path)
        except Exception as e:
            logging.warning(f"LocalWriter: xlsx write failed: {e}")

    def write_cycle(
        self,
        signal: Optional[dict],
        top_n: list,
        systems: list,
        regime_label: Optional[str] = None,
        ts: Optional[datetime] = None,
        candle_cache: Optional[dict] = None,
    ) -> None:
        """Convenience: write Current xlsx + append Log row in one call."""
        ts = ts or datetime.now(timezone.utc)
        self.update_price_tracker(top_n, ts)
        self.write_current_xlsx(signal, top_n, systems, regime_label, ts)
        self.append_log_row(signal, systems, regime_label, ts)
        self.append_ranked_log(top_n, ts)
        self.write_snapshot(top_n, ts)
        if candle_cache is not None:
            self.append_ohlc_log(top_n, candle_cache, ts)
