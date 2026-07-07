"""
signal_tracker.py — Signal Performance Tracker for the Normalized Engine

Runs after engine_normalized.py. On each run it:
  1. Reads the latest normalized engine xlsx output
  2. Logs any new top signals (ticker, timeframe, outfit, rank, entry price, timestamp)
  3. Fetches the current price for every previously logged signal via Webull
  4. Calculates return % and rank change since initial scan
  5. Saves a running performance xlsx and updates the persistent signal log

Persistent log: /cache/output/normalized_engine/signal_log.json
Performance xlsx: /cache/output/normalized_engine/signal_performance_YYYY-MM-DD.xlsx

Usage:
    docker exec e47_engine python /app/signal_tracker.py
    docker exec e47_engine python /app/signal_tracker.py --top 50
    docker exec e47_engine python /app/signal_tracker.py --top 100 --days 30
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import sys
import logging
from datetime import datetime, timezone, timedelta
from pathlib import Path

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

# ── Load .env ─────────────────────────────────────────────────────────────────
_env_path = Path(__file__).parent / ".env"
if _env_path.exists():
    with open(_env_path) as _f:
        for _line in _f:
            _line = _line.strip()
            if _line and not _line.startswith("#") and "=" in _line:
                _k, _, _v = _line.partition("=")
                os.environ.setdefault(_k.strip(), _v.strip())

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"ERROR: Missing dependency — {e}")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
OUTPUT_DIR   = Path(__file__).parent / "output" / "normalized_engine"
SIGNAL_LOG   = OUTPUT_DIR / "signal_log.json"

# ── Colour palette ────────────────────────────────────────────────────────────
COL_HEADER   = "1F4E79"
COL_POS_HI   = "1B5E20"   # return > +5%
COL_POS_MED  = "43A047"   # return +1% to +5%
COL_POS_LOW  = "A5D6A7"   # return 0% to +1%
COL_FLAT     = "FFF9C4"   # return -1% to 0%
COL_NEG_LOW  = "EF9A9A"   # return -1% to -5%
COL_NEG_HI   = "B71C1C"   # return < -5%
COL_NEW      = "E3F2FD"   # new entry this run


def load_signal_log() -> dict:
    """Load the persistent signal log from disk."""
    if SIGNAL_LOG.exists():
        with open(SIGNAL_LOG) as f:
            return json.load(f)
    return {}


def save_signal_log(log: dict) -> None:
    """Save the persistent signal log to disk."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    with open(SIGNAL_LOG, "w") as f:
        json.dump(log, f, indent=2, default=str)
    logging.info(f"Signal log saved: {SIGNAL_LOG} ({len(log)} signals)")


def load_latest_normalized_xlsx() -> pd.DataFrame | None:
    """Load the most recent normalized engine xlsx output."""
    files = sorted(glob.glob(str(OUTPUT_DIR / "normalized_*.xlsx")))
    if not files:
        logging.error(f"No normalized_*.xlsx found in {OUTPUT_DIR}")
        return None
    latest = files[-1]
    logging.info(f"Loading: {latest}")
    df = pd.read_excel(latest, engine="openpyxl")
    df.columns = [c.strip() for c in df.columns]
    return df


def fetch_current_prices(tickers: list[str]) -> dict[str, float]:
    """
    Fetch the latest close price for each ticker via Webull.
    Returns {ticker: price}. Falls back to 0.0 on any error.
    """
    try:
        from engine_normalized import WebullClient
    except ImportError:
        try:
            from engine import WebullClient
        except ImportError:
            logging.warning("Could not import WebullClient — prices will be unavailable")
            return {}

    app_key    = os.environ.get("WEBULL_APP_KEY", "")
    app_secret = os.environ.get("WEBULL_APP_SECRET", "")
    if not (app_key and app_secret):
        logging.warning("WEBULL_APP_KEY/WEBULL_APP_SECRET not set — prices unavailable")
        return {}

    client = WebullClient(app_key, app_secret,
                          region=os.environ.get("WEBULL_REGION", "us"))
    prices = {}
    for ticker in tickers:
        try:
            df = client.fetch_bars(ticker, "1d", count=2)
            if not df.empty:
                prices[ticker] = float(df["close"].iloc[-1])
            else:
                prices[ticker] = 0.0
        except Exception as e:
            logging.debug(f"Price fetch failed for {ticker}: {e}")
            prices[ticker] = 0.0
    return prices


def return_color(ret_pct: float | None) -> str:
    """Return hex fill color based on return %."""
    if ret_pct is None:
        return "F5F5F5"
    if ret_pct >= 5.0:   return COL_POS_HI
    if ret_pct >= 1.0:   return COL_POS_MED
    if ret_pct >= 0.0:   return COL_POS_LOW
    if ret_pct >= -1.0:  return COL_FLAT
    if ret_pct >= -5.0:  return COL_NEG_LOW
    return COL_NEG_HI


def return_font_color(ret_pct: float | None) -> str:
    if ret_pct is None:
        return "AAAAAA"
    if ret_pct >= 5.0 or ret_pct < -5.0:
        return "FFFFFF"
    return "1A1A1A"


def rank_change_str(initial_rank: int, current_rank: int | None) -> str:
    if current_rank is None:
        return "—"
    diff = initial_rank - current_rank  # positive = improved
    if diff > 0:   return f"↑{diff}"
    if diff < 0:   return f"↓{abs(diff)}"
    return "→"


def build_performance_xlsx(log: dict, current_prices: dict,
                           latest_df: pd.DataFrame,
                           out_path: str, top: int) -> None:
    """
    Build the performance tracking spreadsheet.

    Columns:
      Ticker | TF | Outfit | Outfit Periods | Initial Rank | Initial Price |
      Initial Date | Current Price | Return % | Days Held | Current Rank |
      Rank Change | Convergence | Norm Score (initial)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Signal Performance"

    thin  = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:N1")
    title = ws["A1"]
    title.value = f"NORMALIZED ENGINE — Signal Performance Tracker  |  Run: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    title.font  = Font(bold=True, size=13, color="FFFFFF")
    title.fill  = PatternFill("solid", fgColor=COL_HEADER)
    title.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Column headers ────────────────────────────────────────────────────────
    headers = [
        "Ticker", "Timeframe", "Outfit ID", "Outfit Name", "Outfit Periods",
        "Initial Rank", "Initial Price", "Initial Date",
        "Current Price", "Return %", "Days Held",
        "Current Rank", "Rank Change", "Conv (initial)"
    ]
    header_fill = PatternFill("solid", fgColor=COL_HEADER)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.fill      = header_fill
        c.font      = header_font
        c.alignment = center
    ws.row_dimensions[2].height = 20

    # ── Build current rank lookup from latest normalized run ──────────────────
    current_rank_lookup: dict[str, int] = {}
    for _, row in latest_df.iterrows():
        key = f"{row.get('Ticker','').strip()}|{row.get('Timeframe','').strip()}|{row.get('Outfit ID','')}"
        current_rank_lookup[key] = int(row.get("Rank", 0))

    # ── Sort log by initial rank ───────────────────────────────────────────────
    sorted_signals = sorted(log.values(), key=lambda x: x.get("initial_rank", 9999))[:top]

    now = datetime.now(timezone.utc)

    for row_idx, sig in enumerate(sorted_signals, 3):
        ws.row_dimensions[row_idx].height = 18

        ticker      = sig.get("ticker", "")
        timeframe   = sig.get("timeframe", "")
        outfit_id   = sig.get("outfit_id", "")
        init_rank   = sig.get("initial_rank")
        init_price  = sig.get("initial_price", 0.0)
        init_ts_str = sig.get("initial_ts", "")
        convergence = sig.get("convergence", "")
        norm_score  = sig.get("norm_score", "")
        is_new      = sig.get("is_new", False)

        # Current price & return
        curr_price  = current_prices.get(ticker, 0.0)
        ret_pct     = ((curr_price - init_price) / init_price * 100) if init_price and curr_price else None

        # Days held
        try:
            init_dt   = datetime.fromisoformat(init_ts_str.replace("Z", "+00:00"))
            days_held = (now - init_dt).days
        except Exception:
            days_held = "—"

        # Current rank
        sig_key      = f"{ticker}|{timeframe}|{outfit_id}"
        current_rank = current_rank_lookup.get(sig_key)

        # Row fill
        if is_new:
            row_fill = PatternFill("solid", fgColor=COL_NEW)
        else:
            row_fill = PatternFill("solid", fgColor="FFFFFF" if row_idx % 2 == 0 else "F8F9FA")

        base_font = Font(size=10)

        values = [
            ticker,
            timeframe,
            outfit_id,
            sig.get("outfit_name", ""),
            sig.get("outfit_periods", ""),
            init_rank,
            round(init_price, 2) if init_price else "—",
            init_ts_str[:10] if init_ts_str else "—",
            round(curr_price, 2) if curr_price else "—",
            round(ret_pct, 2) if ret_pct is not None else "—",
            days_held,
            current_rank if current_rank else "—",
            rank_change_str(init_rank, current_rank) if init_rank else "—",
            convergence,
        ]

        for col, val in enumerate(values, 1):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.border    = border
            c.alignment = center
            c.font      = base_font

            # Colour the Return % cell
            if col == 10 and ret_pct is not None:
                c.fill = PatternFill("solid", fgColor=return_color(ret_pct))
                c.font = Font(size=10, bold=True,
                              color=return_font_color(ret_pct))
            elif not is_new:
                c.fill = row_fill

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = [8, 10, 9, 26, 30, 12, 14, 13, 14, 10, 10, 13, 12, 14]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A3"

    # ── Summary tab ───────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.merge_cells("A1:F1")
    s_title = ws2["A1"]
    s_title.value = "Performance Summary"
    s_title.font  = Font(bold=True, size=13, color="FFFFFF")
    s_title.fill  = PatternFill("solid", fgColor=COL_HEADER)
    s_title.alignment = center
    ws2.row_dimensions[1].height = 26

    stats_headers = ["Metric", "Value"]
    for col, h in enumerate(stats_headers, 1):
        c = ws2.cell(row=2, column=col, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = center

    rets = [
        ((current_prices.get(s["ticker"], 0) - s["initial_price"]) / s["initial_price"] * 100)
        for s in sorted_signals
        if s.get("initial_price") and current_prices.get(s["ticker"])
    ]

    winners   = [r for r in rets if r > 0]
    losers    = [r for r in rets if r < 0]
    avg_ret   = sum(rets) / len(rets) if rets else 0
    best      = max(rets) if rets else 0
    worst     = min(rets) if rets else 0
    win_rate  = len(winners) / len(rets) * 100 if rets else 0

    stats = [
        ("Total signals tracked", len(sorted_signals)),
        ("Signals with price data", len(rets)),
        ("Winners (return > 0)", len(winners)),
        ("Losers (return < 0)", len(losers)),
        ("Win rate", f"{win_rate:.1f}%"),
        ("Average return", f"{avg_ret:.2f}%"),
        ("Best return", f"{best:.2f}%"),
        ("Worst return", f"{worst:.2f}%"),
        ("New signals this run", sum(1 for s in sorted_signals if s.get("is_new"))),
        ("Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
    ]

    for row_idx, (metric, value) in enumerate(stats, 3):
        ws2.cell(row=row_idx, column=1, value=metric).font = Font(bold=True, size=10)
        c = ws2.cell(row=row_idx, column=2, value=value)
        c.font = Font(size=10)
        c.alignment = center

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 20

    wb.save(out_path)
    print(f"\n  ✅ Performance report saved: {out_path}")


def main():
    parser = argparse.ArgumentParser(description="Normalized Engine Signal Performance Tracker")
    parser.add_argument("--top",  type=int, default=50,
                        help="Top N signals to track from each run (default: 50)")
    parser.add_argument("--days", type=int, default=90,
                        help="Drop signals older than N days (default: 90)")
    args = parser.parse_args()

    print("\n" + "═" * 71)
    print("  SIGNAL TRACKER — Normalized Engine")
    print("═" * 71)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Load existing log ─────────────────────────────────────────────────────
    log = load_signal_log()
    print(f"  Existing signals in log: {len(log)}")

    # ── Load latest normalized output ─────────────────────────────────────────
    latest_df = load_latest_normalized_xlsx()
    if latest_df is None:
        print("  ERROR: Run engine_normalized.py --xlsx first")
        return

    print(f"  Latest run: {len(latest_df)} signals")

    # ── Add new top signals to log ────────────────────────────────────────────
    now_ts   = datetime.now(timezone.utc).isoformat()
    new_count = 0

    for _, row in latest_df.head(args.top).iterrows():
        ticker    = str(row.get("Ticker", "")).strip()
        timeframe = str(row.get("Timeframe", "")).strip()
        outfit_id = row.get("Outfit ID", "")
        sig_key   = f"{ticker}|{timeframe}|{outfit_id}"

        if sig_key not in log:
            log[sig_key] = {
                "ticker":         ticker,
                "timeframe":      timeframe,
                "outfit_id":      outfit_id,
                "outfit_name":    str(row.get("Outfit Name", "")),
                "outfit_periods": str(row.get("Outfit Periods", "")),
                "initial_rank":   int(row.get("Rank", 0)),
                "initial_price":  float(row.get("Entry Price", 0) or 0),
                "initial_ts":     now_ts,
                "convergence":    str(row.get("Convergence", "")),
                "norm_score":     float(row.get("Norm Score", 0) or 0),
                "is_new":         True,
                "price_history":  [],
            }
            new_count += 1
        else:
            log[sig_key]["is_new"] = False

    print(f"  New signals added: {new_count}")

    # ── Prune signals older than --days ───────────────────────────────────────
    cutoff = datetime.now(timezone.utc) - timedelta(days=args.days)
    before = len(log)
    log = {
        k: v for k, v in log.items()
        if datetime.fromisoformat(
            v.get("initial_ts", now_ts).replace("Z", "+00:00")
        ) > cutoff
    }
    pruned = before - len(log)
    if pruned:
        print(f"  Pruned {pruned} signals older than {args.days} days")

    # ── Fetch current prices ──────────────────────────────────────────────────
    tickers = list({v["ticker"] for v in log.values()})
    print(f"  Fetching current prices for {len(tickers)} tickers...")
    current_prices = fetch_current_prices(tickers)
    fetched = sum(1 for p in current_prices.values() if p > 0)
    print(f"  Prices fetched: {fetched}/{len(tickers)}")

    # ── Append to price history in log ───────────────────────────────────────
    for sig_key, sig in log.items():
        price = current_prices.get(sig["ticker"], 0.0)
        if price > 0:
            sig["price_history"].append({
                "ts":    now_ts,
                "price": price,
            })
            # Keep last 180 price snapshots
            sig["price_history"] = sig["price_history"][-180:]

    # ── Save updated log ──────────────────────────────────────────────────────
    save_signal_log(log)

    # ── Build performance xlsx ────────────────────────────────────────────────
    ts       = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M-%S")
    out_path = str(OUTPUT_DIR / f"signal_performance_{ts}.xlsx")
    build_performance_xlsx(log, current_prices, latest_df, out_path, top=args.top)

    # ── Print quick summary to terminal ───────────────────────────────────────
    signals_with_prices = [
        (v["ticker"], v["timeframe"],
         v.get("initial_price", 0), current_prices.get(v["ticker"], 0))
        for v in sorted(log.values(), key=lambda x: x.get("initial_rank", 9999))[:20]
        if current_prices.get(v["ticker"], 0) > 0
    ]

    if signals_with_prices:
        print(f"\n  {'#':<4} {'Ticker':<8} {'TF':<6} {'Entry':>8} {'Now':>8} {'Return':>8}")
        print("  " + "─" * 48)
        for i, (ticker, tf, entry, now_price) in enumerate(signals_with_prices, 1):
            ret = (now_price - entry) / entry * 100 if entry else 0
            arrow = "↑" if ret > 0 else ("↓" if ret < 0 else "→")
            print(f"  {i:<4} {ticker:<8} {tf:<6} {entry:>8.2f} {now_price:>8.2f} {arrow}{abs(ret):>6.2f}%")

    print("\n" + "═" * 71 + "\n")


if __name__ == "__main__":
    main()
