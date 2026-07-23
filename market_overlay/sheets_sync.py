"""
sheets_sync.py — Writes all engine output categories to Google Sheets.
=====================================================================
Tabs written (latest data — overwrite each sync):
  Current       — main engine snapshot (read from signals_current.xlsx)
  Discovery     — latest discovery engine output
  Confluence    — latest confluence engine output
  Backtest      — latest backtest results
  Triangulation — top triangulated signals (overlay composite)
  Overlay       — The System + Zero Gamma snapshot

Log tabs (append only — historical accumulation, never wiped):
  Snapshot_Log    — every snapshot run, date-stamped
  Confluence_Log  — every confluence run, date-stamped
  Backtest_Log    — every backtest run, date-stamped
  Trades_Log      — every trade suggestion, date-stamped
  Signal_Log      — triangulated signal tracker performance

Run standalone:  python3 sheets_sync.py
Or import and call sync_all() from overlay.py after each refresh.

Never modifies the original engine code or sheets_writer.py.
"""

import os
import sys
import csv
import json
import logging
from datetime import datetime, timezone
from pathlib import Path

# Load .env from parent folder
_env_path = Path(__file__).parent.parent / ".env"
if _env_path.exists():
    for line in _env_path.read_text().splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, _, v = line.partition("=")
            os.environ.setdefault(k.strip(), v.strip())

OUTPUT_DIR   = Path(__file__).parent.parent / "output"
CREDS_PATH   = os.environ.get("GOOGLE_SHEETS_CREDENTIALS_PATH", "")
SHEET_ID     = os.environ.get("GOOGLE_SHEET_ID", "")
SCOPES       = ["https://www.googleapis.com/auth/spreadsheets"]

# State file — tracks last-logged filename per category to avoid re-appending
STATE_FILE   = Path(__file__).parent.parent / "logs" / "sheets_sync_state.json"
LOG_TABS     = ["Snapshot_Log", "Confluence_Log", "Backtest_Log", "Trades_Log",
                "Signal_Log", "V2_Signal_Log", "Main_Signal_Log",
                "Current_Log", "Discovery_Log", "Normalized_Log", "V3_Log",
                "Triangulation_Log", "Overlay_Log", "Operator_Log"]

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s")
log = logging.getLogger("sheets_sync")


# ── Google Sheets client ──────────────────────────────────────────────────────

def _build_service():
    try:
        from googleapiclient.discovery import build
        from google.oauth2 import service_account
        creds = service_account.Credentials.from_service_account_file(
            CREDS_PATH, scopes=SCOPES)
        svc = build("sheets", "v4", credentials=creds, cache_discovery=False)
        # Quick connection test
        svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
        return svc
    except Exception as e:
        log.warning(f"Sheets connection failed: {e}")
        return None


def _ensure_tabs(svc, tab_names: list[str]):
    try:
        meta     = svc.spreadsheets().get(spreadsheetId=SHEET_ID).execute()
        existing = {s["properties"]["title"] for s in meta.get("sheets", [])}
        requests = [{"addSheet": {"properties": {"title": t}}}
                    for t in tab_names if t not in existing]
        if requests:
            svc.spreadsheets().batchUpdate(
                spreadsheetId=SHEET_ID,
                body={"requests": requests}).execute()
            log.info(f"Created tabs: {[r['addSheet']['properties']['title'] for r in requests]}")
    except Exception as e:
        log.warning(f"Tab creation error: {e}")


def _write_tab(svc, tab: str, rows: list[list]):
    """Clear tab and write rows."""
    try:
        svc.spreadsheets().values().clear(
            spreadsheetId=SHEET_ID,
            range=f"{tab}!A1:Z5000").execute()
        if rows:
            svc.spreadsheets().values().update(
                spreadsheetId=SHEET_ID,
                range=f"{tab}!A1",
                valueInputOption="RAW",
                body={"values": rows}).execute()
        log.info(f"  ✓ {tab}: {len(rows)} rows written")
    except Exception as e:
        log.warning(f"  ✗ {tab}: write failed — {e}")


def _append_tab(svc, tab: str, rows: list[list]):
    """Append rows to a log tab — never clears existing data."""
    try:
        svc.spreadsheets().values().append(
            spreadsheetId=SHEET_ID,
            range=f"{tab}!A1",
            valueInputOption="RAW",
            insertDataOption="INSERT_ROWS",
            body={"values": rows}).execute()
        log.info(f"  ✓ {tab}: +{len(rows)} rows appended")
    except Exception as e:
        log.warning(f"  ✗ {tab}: append failed — {e}")


def _load_sync_state() -> dict:
    """Load last-synced file tracking state."""
    if STATE_FILE.exists():
        try:
            return json.loads(STATE_FILE.read_text())
        except Exception:
            return {}
    return {}


def _save_sync_state(state: dict):
    """Persist sync state to disk."""
    STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
    STATE_FILE.write_text(json.dumps(state, indent=2))


def _prepend_date(rows: list[list], date_str: str) -> list[list]:
    """Prepend a date/timestamp column to every row."""
    return [[date_str] + [str(v) for v in row] for row in rows]


# ── Data readers ──────────────────────────────────────────────────────────────

def _latest_csv(folder: str) -> Path | None:
    d = OUTPUT_DIR / folder
    if not d.exists():
        return None
    files = sorted(d.glob("*.csv"))
    return files[-1] if files else None


def _read_csv(path: Path) -> list[list]:
    if not path or not path.exists():
        return []
    rows = []
    with open(path, newline="") as f:
        for row in csv.reader(f):
            rows.append(row)
    return rows


def _latest_root_csv(prefix: str) -> Path | None:
    files = sorted(OUTPUT_DIR.glob(f"{prefix}*.csv"))
    return files[-1] if files else None


# ── Tab builders ──────────────────────────────────────────────────────────────

def build_current_rows() -> list[list]:
    """Read signals_current.xlsx and return all rows for the Current tab."""
    xlsx_path = OUTPUT_DIR / "signals_current.xlsx"
    if not xlsx_path.exists():
        return [["No signals_current.xlsx found"]]
    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([("" if v is None else v) for v in row])
        wb.close()
        return rows
    except Exception as e:
        return [["Error reading signals_current.xlsx"], [str(e)]]


def build_discovery_rows() -> list[list]:
    path = _latest_csv("discovery")
    rows = _read_csv(path)
    if not rows:
        return [["No discovery data found"]]
    header_extra = [["DISCOVERY ENGINE — Latest Output"],
                    [f"File: {path.name if path else '—'}"],
                    [f"Updated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"],
                    []]
    return header_extra + rows


def build_confluence_rows() -> list[list]:
    path = _latest_csv("confluence")
    rows = _read_csv(path)
    if not rows:
        return [["No confluence data found"]]
    header_extra = [["CONFLUENCE ENGINE — Latest Output"],
                    [f"File: {path.name if path else '—'}"],
                    [f"Updated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"],
                    []]
    return header_extra + rows


def build_backtest_rows() -> list[list]:
    path = _latest_root_csv("backtest_")
    rows = _read_csv(path)
    if not rows:
        return [["No backtest data found"]]
    header_extra = [["BACKTEST — Latest Output"],
                    [f"File: {path.name if path else '—'}"],
                    [f"Updated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"],
                    []]
    return header_extra + rows


def build_trades_rows() -> list[list]:
    path = _latest_csv("trades")
    rows = _read_csv(path)
    if not rows:
        # Try root-level trades files
        path = _latest_root_csv("trades_")
        rows = _read_csv(path)
    if not rows:
        return [["No trade signals found — run trade engine first"]]
    header_extra = [["TRADE ENGINE — Latest Output"],
                    [f"File: {path.name if path else '—'}"],
                    [f"Updated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"],
                    []]
    return header_extra + rows


def build_triangulation_rows() -> list[list]:
    """Read triangulation output from the _triangulation_staging folder."""
    try:
        tri_path = Path(__file__).parent.parent / "_triangulation_staging"
        sys.path.insert(0, str(tri_path))
        from triangulator import (
            read_original, read_normalized, read_v3,
            read_discovery, read_confluence, read_trades,
            triangulate,
        )
        orig  = read_original()
        norm  = read_normalized()
        v3    = read_v3()
        disc  = read_discovery()
        conf  = read_confluence()
        trade = read_trades()
        signals = triangulate(orig, norm, v3, disc, conf, trade)

        rows = [
            ["TRIANGULATION — Top Signals"],
            [f"Updated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"],
            [],
            ["#", "Ticker", "Score", "Engines", "V3 State", "Norm Conv",
             "Orig Conv", "Disc TF", "Disc SMA", "Entry Price"],
        ]
        shown = [s for s in signals if s.score > 0]
        for i, sig in enumerate(shown, 1):
            rows.append([
                i,
                sig.ticker,
                round(sig.score, 1),
                f"{sig.engines_confirmed}/3",
                sig.v3_state or "—",
                sig.norm_conv or "—",
                sig.orig_conv or "—",
                sig.disc_tf or "—",
                sig.disc_sma or "—",
                f"${sig.trade_entry:.2f}" if sig.trade_entry else "—",
            ])
        return rows
    except Exception as e:
        return [["Triangulation data unavailable"], [str(e)]]


def build_v3_rows() -> list[list]:
    import openpyxl
    v3_dir = OUTPUT_DIR / "v3"
    path = None
    if v3_dir.exists():
        xlsx_files = sorted(v3_dir.glob("*.xlsx"))
        if xlsx_files:
            path = xlsx_files[-1]
    if not path:
        return [["No V3 data found"]]
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [[("" if v is None else str(v)) for v in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
    except Exception as e:
        return [[f"V3 read error: {e}"]]
    if not rows:
        return [["No V3 data found"]]
    header_extra = [["V3 ENGINE — Latest Output"],
                    [f"File: {path.name}"],
                    [f"Updated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"],
                    []]
    return header_extra + rows


def build_operator_rows() -> list[list]:
    """Latest operator board export (opboard.py --export)."""
    import openpyxl
    op_dir = OUTPUT_DIR / "operator"
    path = None
    if op_dir.exists():
        xlsx_files = sorted(op_dir.glob("operator_*.xlsx"))
        if xlsx_files:
            path = xlsx_files[-1]
    if not path:
        return [["No operator data found — run: python3 opboard.py --export"]]
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = [[("" if v is None else str(v)) for v in row]
                for row in ws.iter_rows(values_only=True)]
        wb.close()
    except Exception as e:
        return [[f"Operator read error: {e}"]]
    if not rows:
        return [["No operator data found"]]
    header_extra = [["OPERATOR BOARD — Raw Cross-Engine View"],
                    [f"File: {path.name}"],
                    [f"Updated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"],
                    []]
    return header_extra + rows


def build_normalized_rows() -> list[list]:
    path = _latest_csv("normalized")
    rows = _read_csv(path)
    if not rows:
        norm_dir = OUTPUT_DIR / "normalized_engine"
        if norm_dir.exists():
            files = sorted(norm_dir.glob("*.csv"))
            if files:
                path = files[-1]
                rows = _read_csv(path)
    if not rows:
        return [["No normalized engine data found"]]
    header_extra = [["NORMALIZED ENGINE — Latest Output"],
                    [f"File: {path.name if path else '—'}"],
                    [f"Updated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"],
                    []]
    return header_extra + rows


def build_overlay_rows(sys_data: dict = None, gex_data: dict = None) -> list[list]:
    """Snapshot of The System + Zero Gamma for the Overlay tab."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    rows = [
        ["MARKET OVERLAY SNAPSHOT"],
        [f"Updated: {now}"],
        [],
    ]
    if sys_data and "error" not in sys_data:
        rows += [
            ["THE SYSTEM"],
            ["State",        sys_data.get("state", "—")],
            ["Vehicle",      sys_data.get("vehicle", "—")],
            ["SPY Close",    sys_data.get("close", "—")],
            ["SMA10",        sys_data.get("sma10", "—")],
            ["SMA50",        sys_data.get("sma50", "—")],
            ["SMA200",       sys_data.get("sma200", "—")],
            ["SMA50 Dir",    sys_data.get("sma50_dir", "—")],
            ["Condition",    sys_data.get("ob_os", "—")],
            ["Signal",       sys_data.get("entry_signal", "—")],
            ["Reason",       sys_data.get("entry_reason", "—")],
            ["Choppy",       sys_data.get("choppy", False)],
            [],
        ]
        ndx = sys_data.get("nasdaq", {})
        if ndx and "error" not in ndx:
            rows += [
                ["NASDAQ (QQQ)"],
                ["State",     ndx.get("state", "—")],
                ["QQQ Close", ndx.get("close", "—")],
                ["Day %",     f"{ndx.get('gap_pct', 0):+.2f}%"],
                ["vs SMA50",  f"{ndx.get('dist_sma50', 0):+.2f}%"],
                ["SMA50 Dir", ndx.get("sma50_dir", "—")],
                [],
            ]
    if gex_data and "error" not in gex_data:
        rows += [
            ["ZERO GAMMA (SPX — Tikitrade)"],
            ["Spot",         gex_data.get("spot", "—")],
            ["Zero Gamma",   gex_data.get("zero_gamma", "—")],
            ["Regime",       gex_data.get("regime", "—")],
            ["Distance %",   f"{gex_data.get('dist_from_zero_pct', 0):+.3f}%"],
            ["Call Wall",    gex_data.get("call_wall", "—")],
            ["Put Wall",     gex_data.get("put_wall", "—")],
            ["Max Pain",     gex_data.get("max_pain", "—")],
            ["Data As Of",   gex_data.get("data_as_of", "—")],
        ]
    return rows


# ── Log tab builders (append-only, historical accumulation) ──────────────────

def _sync_logs(svc, state: dict) -> dict:
    """
    Append new rows to each log tab if the underlying file has changed since
    last sync. Returns updated state dict.
    """
    now_str = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    # ── Snapshot_Log ──────────────────────────────────────────────────────────
    snap_files = sorted((OUTPUT_DIR / "snapshots").glob("*.csv")) if (OUTPUT_DIR / "snapshots").exists() else []
    if snap_files:
        snap_path = snap_files[-1]
        snap_key  = f"snapshot:{snap_path.name}"
        if state.get("Snapshot_Log") != snap_key:
            rows = _read_csv(snap_path)
            if rows:
                header = rows[0]
                data   = rows[1:]
                if not state.get("Snapshot_Log"):          # first time — write header
                    _append_tab(svc, "Snapshot_Log", [["synced_utc"] + header])
                _append_tab(svc, "Snapshot_Log", _prepend_date(data, now_str))
                state["Snapshot_Log"] = snap_key

    # ── Confluence_Log ────────────────────────────────────────────────────────
    conf_path = _latest_csv("confluence")
    if conf_path:
        conf_key = f"confluence:{conf_path.name}"
        if state.get("Confluence_Log") != conf_key:
            rows = _read_csv(conf_path)
            if rows:
                header = rows[0]
                data   = rows[1:]
                if not state.get("Confluence_Log"):
                    _append_tab(svc, "Confluence_Log", [["synced_utc"] + header])
                _append_tab(svc, "Confluence_Log", _prepend_date(data, now_str))
                state["Confluence_Log"] = conf_key

    # ── Backtest_Log ──────────────────────────────────────────────────────────
    bt_path = _latest_root_csv("backtest_")
    if bt_path:
        bt_key = f"backtest:{bt_path.name}"
        if state.get("Backtest_Log") != bt_key:
            rows = _read_csv(bt_path)
            if rows:
                header = rows[0]
                data   = rows[1:]
                if not state.get("Backtest_Log"):
                    _append_tab(svc, "Backtest_Log", [["synced_utc"] + header])
                _append_tab(svc, "Backtest_Log", _prepend_date(data, now_str))
                state["Backtest_Log"] = bt_key

    # ── Trades_Log ────────────────────────────────────────────────────────────
    trades_path = _latest_csv("trades")
    if not trades_path:
        trades_path = _latest_root_csv("trades_")
    if trades_path:
        trades_key = f"trades:{trades_path.name}"
        if state.get("Trades_Log") != trades_key:
            rows = _read_csv(trades_path)
            if rows:
                header = rows[0]
                data   = rows[1:]
                if not state.get("Trades_Log"):
                    _append_tab(svc, "Trades_Log", [["synced_utc"] + header])
                _append_tab(svc, "Trades_Log", _prepend_date(data, now_str))
                state["Trades_Log"] = trades_key

    # ── Signal_Log (triangulated tracker) ────────────────────────────────────
    sig_dir = OUTPUT_DIR / "signal_tracking"
    if sig_dir.exists():
        sig_files = sorted(sig_dir.glob("triangulated_performance_*.csv"))
        if sig_files:
            sig_path = sig_files[-1]
            sig_key  = f"signal:{sig_path.name}"
            if state.get("Signal_Log") != sig_key:
                rows = _read_csv(sig_path)
                if rows:
                    header = rows[0]
                    data   = rows[1:]
                    if not state.get("Signal_Log"):
                        _append_tab(svc, "Signal_Log", [["synced_utc"] + header])
                    _append_tab(svc, "Signal_Log", _prepend_date(data, now_str))
                    state["Signal_Log"] = sig_key

    # ── V2_Signal_Log (V3 engine signal tracker) ──────────────────────────────
    if sig_dir.exists():
        v2_files = sorted(sig_dir.glob("performance_*.csv"))
        if v2_files:
            v2_path = v2_files[-1]
            v2_key  = f"v2signal:{v2_path.name}"
            if state.get("V2_Signal_Log") != v2_key:
                rows = _read_csv(v2_path)
                if rows:
                    header = rows[0]
                    data   = rows[1:]
                    if not state.get("V2_Signal_Log"):
                        _append_tab(svc, "V2_Signal_Log", [["synced_utc"] + header])
                    _append_tab(svc, "V2_Signal_Log", _prepend_date(data, now_str))
                    state["V2_Signal_Log"] = v2_key

    # ── Main_Signal_Log (main engine signal tracker) ──────────────────────────
    if sig_dir.exists():
        main_files = sorted(sig_dir.glob("main_performance_*.csv"))
        if main_files:
            main_path = main_files[-1]
            main_key  = f"mainsignal:{main_path.name}"
            if state.get("Main_Signal_Log") != main_key:
                rows = _read_csv(main_path)
                if rows:
                    header = rows[0]
                    data   = rows[1:]
                    if not state.get("Main_Signal_Log"):
                        _append_tab(svc, "Main_Signal_Log", [["synced_utc"] + header])
                    _append_tab(svc, "Main_Signal_Log", _prepend_date(data, now_str))
                    state["Main_Signal_Log"] = main_key

    # ── Current_Log ───────────────────────────────────────────────────────────
    current_path = OUTPUT_DIR / "signals_current.xlsx"
    if current_path.exists():
        current_key = f"current:{current_path.stat().st_mtime:.0f}"
        if state.get("Current_Log") != current_key:
            rows = build_current_rows()
            if rows:
                if not state.get("Current_Log"):
                    _append_tab(svc, "Current_Log", [["synced_utc"] + list(rows[0])])
                _append_tab(svc, "Current_Log", _prepend_date(rows[1:], now_str))
            state["Current_Log"] = current_key

    # ── Discovery_Log ─────────────────────────────────────────────────────────
    disc_log_path = _latest_csv("discovery")
    if disc_log_path:
        disc_log_key = f"discovery_log:{disc_log_path.name}"
        if state.get("Discovery_Log") != disc_log_key:
            rows = _read_csv(disc_log_path)
            if rows:
                if not state.get("Discovery_Log"):
                    _append_tab(svc, "Discovery_Log", [["synced_utc"] + rows[0]])
                _append_tab(svc, "Discovery_Log", _prepend_date(rows[1:], now_str))
            state["Discovery_Log"] = disc_log_key

    # ── Normalized_Log ────────────────────────────────────────────────────────
    norm_log_dir = OUTPUT_DIR / "normalized_engine"
    if norm_log_dir.exists():
        norm_log_files = sorted(norm_log_dir.glob("*.xlsx"))
        if norm_log_files:
            norm_log_path = norm_log_files[-1]
            norm_log_key = f"normalized_log:{norm_log_path.name}"
            if state.get("Normalized_Log") != norm_log_key:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(norm_log_path, read_only=True, data_only=True)
                    ws = wb.active
                    rows = [[("" if v is None else v) for v in row] for row in ws.iter_rows(values_only=True)]
                    wb.close()
                    if rows:
                        if not state.get("Normalized_Log"):
                            _append_tab(svc, "Normalized_Log", [["synced_utc"] + list(rows[0])])
                        _append_tab(svc, "Normalized_Log", _prepend_date(rows[1:], now_str))
                    state["Normalized_Log"] = norm_log_key
                except Exception as e:
                    log.warning(f"Normalized_Log read error: {e}")

    # ── V3_Log ────────────────────────────────────────────────────────────────
    v3_log_dir = OUTPUT_DIR / "v3"
    if v3_log_dir.exists():
        v3_log_files = sorted(v3_log_dir.glob("*.xlsx"))
        if v3_log_files:
            v3_log_path = v3_log_files[-1]
            v3_log_key = f"v3_log:{v3_log_path.name}"
            if state.get("V3_Log") != v3_log_key:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(v3_log_path, read_only=True, data_only=True)
                    ws = wb.active
                    rows = [[("" if v is None else v) for v in row] for row in ws.iter_rows(values_only=True)]
                    wb.close()
                    if rows:
                        if not state.get("V3_Log"):
                            _append_tab(svc, "V3_Log", [["synced_utc"] + list(rows[0])])
                        _append_tab(svc, "V3_Log", _prepend_date(rows[1:], now_str))
                    state["V3_Log"] = v3_log_key
                except Exception as e:
                    log.warning(f"V3_Log read error: {e}")

    # ── Operator_Log ──────────────────────────────────────────────────────────
    op_dir = OUTPUT_DIR / "operator"
    if op_dir.exists():
        op_files = sorted(op_dir.glob("operator_*.xlsx"))
        if op_files:
            op_path = op_files[-1]
            op_key = f"operator_log:{op_path.name}"
            if state.get("Operator_Log") != op_key:
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(op_path, read_only=True, data_only=True)
                    ws = wb.active
                    rows = [[("" if v is None else v) for v in row]
                            for row in ws.iter_rows(values_only=True)]
                    wb.close()
                    if rows:
                        if not state.get("Operator_Log"):
                            _append_tab(svc, "Operator_Log", [["synced_utc"] + list(rows[0])])
                        _append_tab(svc, "Operator_Log", _prepend_date(rows[1:], now_str))
                    state["Operator_Log"] = op_key
                except Exception as e:
                    log.warning(f"Operator_Log read error: {e}")

    # ── Triangulation_Log — append once per V3 cycle ──────────────────────────
    v3_key_for_tri = state.get("V3_Log", "")
    if v3_key_for_tri and state.get("Triangulation_Log") != v3_key_for_tri:
        try:
            tri_rows = build_triangulation_rows()
            data_rows = [r for r in tri_rows if r and len(r) > 1 and r[0] not in
                         ["TRIANGULATION — Top Signals", "Updated:", ""]]
            if data_rows:
                if not state.get("Triangulation_Log"):
                    _append_tab(svc, "Triangulation_Log",
                                [["synced_utc", "#", "Ticker", "Score", "Engines",
                                  "V3 State", "Norm Conv", "Orig Conv", "Disc TF",
                                  "Disc SMA", "Entry Price"]])
                _append_tab(svc, "Triangulation_Log", _prepend_date(data_rows, now_str))
            state["Triangulation_Log"] = v3_key_for_tri
        except Exception as e:
            log.warning(f"Triangulation_Log error: {e}")

    # ── Overlay_Log — append once per V3 cycle ───────────────────────────────
    if v3_key_for_tri and state.get("Overlay_Log") != v3_key_for_tri:
        try:
            ov_rows = build_overlay_rows()
            flat = [r for r in ov_rows if r and any(str(v).strip() for v in r)]
            if flat:
                if not state.get("Overlay_Log"):
                    _append_tab(svc, "Overlay_Log", [["synced_utc", "key", "value"]])
                _append_tab(svc, "Overlay_Log", _prepend_date(flat, now_str))
            state["Overlay_Log"] = v3_key_for_tri
        except Exception as e:
            log.warning(f"Overlay_Log error: {e}")

    return state


# ── Main sync ─────────────────────────────────────────────────────────────────

TABS = ["Current", "Discovery", "Confluence", "Backtest", "Trades",
        "Normalized", "V3", "Triangulation", "Overlay", "Operator"] + LOG_TABS


def sync_all(sys_data: dict = None, gex_data: dict = None):
    """Sync all tabs. Call from overlay after each refresh, or run standalone."""
    if not CREDS_PATH or not SHEET_ID:
        log.warning("GOOGLE_SHEETS_CREDENTIALS_PATH or GOOGLE_SHEET_ID not set — skipping sync")
        return

    svc = _build_service()
    if not svc:
        return

    _ensure_tabs(svc, TABS)

    tab_data = {
        "Current":       build_current_rows,
        "Discovery":     build_discovery_rows,
        "Confluence":    build_confluence_rows,
        "Backtest":      build_backtest_rows,
        "Trades":        build_trades_rows,
        "Normalized":    build_normalized_rows,
        "V3":            build_v3_rows,
        "Triangulation": build_triangulation_rows,
        "Overlay":       lambda: build_overlay_rows(sys_data, gex_data),
        "Operator":      build_operator_rows,
    }

    for tab, builder in tab_data.items():
        try:
            rows = builder()
            _write_tab(svc, tab, rows)
        except Exception as e:
            log.warning(f"  ✗ {tab}: builder error — {e}")

    # ── Append-only log tabs (historical accumulation) ────────────────────────
    log.info("Syncing log tabs...")
    state = _load_sync_state()
    state = _sync_logs(svc, state)
    _save_sync_state(state)

    log.info("Sheets sync complete.")


if __name__ == "__main__":
    log.info("Running full sheets sync...")
    sync_all()
