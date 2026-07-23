"""
outfit_discovery.py — SMA Outfit Discovery Engine

Searches for NEW SMA period combinations exhibiting institutional gravity
across a curated 30-ticker basket:
  - 11 US sector ETFs
  - 9 international ETFs
  - 10 commodity ETFs

Uses the same touch-detection logic as the main engine but applied to
candidate period combinations (geometric sequences) rather than the fixed 42 outfits.

Promotion threshold: fires on 60%+ of basket with 60%+ win rate at +10 candle window.

Usage (run inside engine container):
    docker exec e47_engine python /app/outfit_discovery.py
    docker exec e47_engine python /app/outfit_discovery.py --timeframes 1d,1w
    docker exec e47_engine python /app/outfit_discovery.py --top 20
    docker exec e47_engine python /app/outfit_discovery.py --tickers XLF,XLE,GLD

Output:
    output/outfit_discovery/outfit_discovery_YYYY-MM-DD_HH-MM-SS.csv
"""

from __future__ import annotations

import csv
import gzip
import pickle
import argparse
import logging
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [outfit_discovery] %(message)s",
    datefmt="%H:%M:%S",
)

# ── Paths ─────────────────────────────────────────────────────────────────────
CACHE_DIR  = "/cache/candle_cache"
OUTPUT_DIR = Path("./output/outfit_discovery")

# ── Default 30-ticker basket ──────────────────────────────────────────────────
DEFAULT_TICKERS = [
    # US Sectors (11)
    "XLF", "XLE", "XLK", "XLB", "XLV", "XLI", "XLY", "XLP", "XLU", "XLRE", "XLC",
    # International (9)
    "EWJ", "EWG", "EWU", "FXI", "EWY", "EWZ", "EWC", "EWA", "INDA",
    # Commodities (10)
    "GLD", "SLV", "BITO", "CORN", "WEAT", "SOYB", "USO", "UNG", "CPER", "BAL",
]

# ── Parameters ────────────────────────────────────────────────────────────────
TOUCH_PCT    = 0.003   # price within 0.3% of SMA = touch
ABSENCE_WIN  = 50      # must not have touched this MA in last N candles
MIN_HITS     = 3       # minimum touch events per ticker to qualify
BREADTH_MIN  = 0.60    # outfit must fire on 60% of basket
WIN_RATE_MIN = 0.60    # 60% win rate at +10 candle to be promoted
FWD_WINDOWS  = [1, 3, 5, 10, 20]

# ── Out-of-sample validation ─────────────────────────────────────────────────
# Touches in the first OOS_FRACTION of each series are "in-sample" (discovery);
# touches in the remainder are "out-of-sample" (validation). An outfit that
# only worked in one old regime shows a collapsed oos win rate and is flagged.
OOS_FRACTION = 0.70
OOS_MIN_N    = 5       # need at least this many oos events to validate
OOS_WIN_MIN  = 55.0    # oos win10 >= this (%) to be marked VALIDATED

# ── Special multipliers (the patterns that dominated the Jul 9 run) ──────────
# Pure 0.1-step grids miss these: phi, sqrt2, sqrt3, phi², sqrt(phi), 3/2.
SPECIAL_MULTS = [1.272, 1.414, 1.5, 1.618, 1.732, 2.0, 2.618]

# ── Fibonacci seed windows (scaled) ──────────────────────────────────────────
FIB = [2, 3, 5, 8, 13, 21, 34, 55, 89, 144, 233, 377, 610, 987]

# ── Novel outfit registry — accumulates across runs ──────────────────────────
REGISTRY_PATH = Path("./output/outfit_discovery/all_novel_outfits.csv")


# ─────────────────────────────────────────────────────────────────────────────
# Sequence families — f(seed) → 6-period outfit
# ─────────────────────────────────────────────────────────────────────────────

def _seq_ratio(seed: int, m: float, n: int = 6) -> tuple[int, ...]:
    """seed × m^i — the seed is the SHORTEST period."""
    return tuple(sorted(set(int(round(seed * (m ** i))) for i in range(n))))


def _seq_ratio_anchor(seed: int, m: float, n: int = 6) -> tuple[int, ...]:
    """seed ÷ m^i — the seed is the LONGEST period (anchor a discovered level
    like MA288 at the top and derive the shorter periods underneath it)."""
    return tuple(sorted(set(int(round(seed / (m ** i))) for i in range(n))))


def _seq_fib(seed: int, n: int = 6) -> tuple[int, ...]:
    """Fibonacci ratio chain from the seed: seed × (1,2,3,5,8,13)."""
    ratios = [1, 2, 3, 5, 8, 13][:n]
    return tuple(sorted(set(int(round(seed * r)) for r in ratios)))


def _seq_fib_anchor(seed: int, n: int = 6) -> tuple[int, ...]:
    """Fibonacci chain with the seed as the LONGEST period: seed × (1,2,3,5,8,13)/13."""
    ratios = [1, 2, 3, 5, 8, 13][:n]
    top = ratios[-1]
    return tuple(sorted(set(int(round(seed * r / top)) for r in ratios)))


SEQUENCE_FAMILIES: dict[str, tuple[str, callable]] = {
    "fib":         ("Fibonacci chain up from seed (s×1,2,3,5,8,13)", _seq_fib),
    "fib-anchor":  ("Fibonacci chain down from seed (seed = longest MA)", _seq_fib_anchor),
    "phi":         ("Golden ratio up (s×1.618^i)", lambda s, n=6: _seq_ratio(s, 1.618, n)),
    "phi-anchor":  ("Golden ratio down (seed = longest MA)", lambda s, n=6: _seq_ratio_anchor(s, 1.618, n)),
    "sqrt2":       ("√2 up (s×1.414^i)", lambda s, n=6: _seq_ratio(s, 1.414, n)),
    "sqrt3":       ("√3 up (s×1.732^i)", lambda s, n=6: _seq_ratio(s, 1.732, n)),
    "double":      ("Doubling (s×2^i)", lambda s, n=6: _seq_ratio(s, 2.0, n)),
    "fifth":       ("Perfect fifth (s×1.5^i)", lambda s, n=6: _seq_ratio(s, 1.5, n)),
}


def seq_custom_fx(expr: str):
    """
    Build a sequence function from a user expression in x (seed) and i (index 0-5).
    Example: "x * 1.618**i"  or  "x + i*x//2"  or  "x * (i+1)**1.5"
    Evaluated with a restricted namespace — math module available.
    """
    import math as _math
    def fn(seed: int, n: int = 6) -> tuple[int, ...]:
        vals = set()
        for i in range(n):
            try:
                v = eval(expr, {"__builtins__": {}},
                         {"x": seed, "i": i, "math": _math})
                vals.add(int(round(float(v))))
            except Exception:
                return ()
        return tuple(sorted(vals))
    return fn


# ─────────────────────────────────────────────────────────────────────────────
# Seed sources
# ─────────────────────────────────────────────────────────────────────────────

def seeds_from_influx_periods(tickers: list[str], top: int = 20,
                              window_days: int = 7) -> list[int]:
    """
    Top SMA periods by 7-day cumulative deciseconds across the basket —
    the levels price has actually spent the most time at, per InfluxDB.
    """
    try:
        import os as _os
        from influxdb_client import InfluxDBClient
        client = InfluxDBClient(
            url=_os.environ.get("INFLUX_URL", "http://influxdb:8086"),
            token=_os.environ.get("INFLUX_TOKEN", "element47-dev-token"),
            org=_os.environ.get("INFLUX_ORG", "element47"),
            timeout=45_000)
        or_chain = " or ".join(f'r.ticker == "{t}"' for t in sorted(set(tickers)))
        flux = f"""
from(bucket: "{_os.environ.get('INFLUX_BUCKET', 'sma_engine')}")
  |> range(start: -{window_days}d)
  |> filter(fn: (r) => r._measurement == "hits" and r._field == "deciseconds")
  |> filter(fn: (r) => {or_chain})
  |> group(columns: ["sma_period"])
  |> sum(column: "_value")
"""
        totals: dict[int, float] = {}
        for tb in client.query_api().query(flux):
            for rec in tb.records:
                try:
                    p = int(rec.values.get("sma_period", 0))
                    totals[p] = totals.get(p, 0) + float(rec.get_value() or 0)
                except (ValueError, TypeError):
                    continue
        client.close()
        ranked = sorted(totals.items(), key=lambda kv: -kv[1])[:top]
        seeds = [p for p, _ in ranked if 2 <= p <= 999]
        logging.info(f"Influx seeds (top cumulative-ds SMA periods): {seeds}")
        return seeds
    except Exception as e:
        logging.warning(f"Influx seed query failed: {e}")
        return []


def seeds_from_entry_prices(top: int = 20) -> list[int]:
    """
    Integer entry prices from the current top signals — the price levels
    themselves become sequence seeds (price→period resonance).
    """
    seeds: list[int] = []
    try:
        import openpyxl
        path = Path("./output/signals_current.xlsx")
        if not path.exists():
            return []
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb["Current"] if "Current" in wb.sheetnames else wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        # Entry prices live in the ohlc log — use ranked table's tickers via
        # ohlc_log for robustness
        import csv as _csv
        ohlc = Path("./output/ohlc_log.csv")
        prices: dict[str, float] = {}
        if ohlc.exists():
            with open(ohlc) as f:
                for row in _csv.DictReader(f):
                    try:
                        prices[row["ticker"]] = float(row["close"])
                    except (ValueError, KeyError):
                        continue
        header_idx = next((i for i, r in enumerate(rows) if r and r[0] == "Rank"), None)
        if header_idx is not None:
            seen = set()
            for row in rows[header_idx + 1:]:
                if not row or row[0] is None or not str(row[0]).isdigit():
                    continue
                t = str(row[1] or "")
                if t in prices and t not in seen:
                    seen.add(t)
                    p = int(round(prices[t]))
                    if 2 <= p <= 999:
                        seeds.append(p)
                if len(seeds) >= top:
                    break
        seeds = list(dict.fromkeys(seeds))
        logging.info(f"Entry-price seeds (top signal closes): {seeds}")
        return seeds
    except Exception as e:
        logging.warning(f"Entry-price seed extraction failed: {e}")
        return []


def seeds_from_registry(top: int = 20) -> list[int]:
    """Base (shortest) periods of past registry winners."""
    out = []
    for periods in load_registry_seeds(max_seeds=top):
        if periods and 2 <= periods[0] <= 999:
            out.append(periods[0])
    return list(dict.fromkeys(out))


# ─────────────────────────────────────────────────────────────────────────────
# Interactive mode
# ─────────────────────────────────────────────────────────────────────────────

def interactive_candidates(tickers: list[str]) -> tuple[list[tuple[int, ...]], str]:
    """
    Menu-driven candidate generation. Returns (candidates, run label).
    """
    print("\n══ OUTFIT DISCOVERY — INTERACTIVE ═══════════════════════════")
    print("\nSequence family:")
    fams = list(SEQUENCE_FAMILIES.items())
    for i, (name, (desc, _)) in enumerate(fams, 1):
        print(f"  {i}) {name:<12} {desc}")
    print(f"  {len(fams)+1}) custom       f(x, i) expression, e.g. x * 1.618**i + i")
    choice = input("\nSelect family [1]: ").strip() or "1"
    try:
        ci = int(choice)
    except ValueError:
        ci = 1
    if ci == len(fams) + 1:
        expr = input("  f(x, i) = ").strip()
        fam_name, fam_fn = f"custom({expr})", seq_custom_fx(expr)
    else:
        fam_name, (_, fam_fn) = fams[max(0, min(ci - 1, len(fams) - 1))]

    print("\nSeed source:")
    print("  1) manual — type integers")
    print("  2) influx — top SMA periods by 7-day cumulative deciseconds")
    print("  3) prices — integer entry prices of current top signals")
    print("  4) registry — base periods of past discovered winners")
    src = input("\nSelect source [1]: ").strip() or "1"

    if src == "2":
        seeds = seeds_from_influx_periods(tickers)
        src_name = "influx-cumds"
    elif src == "3":
        seeds = seeds_from_entry_prices()
        src_name = "entry-prices"
    elif src == "4":
        seeds = seeds_from_registry()
        src_name = "registry"
    else:
        raw = input("  Seeds (comma-separated ints): ").strip()
        seeds = [int(s) for s in raw.split(",") if s.strip().isdigit()]
        src_name = "manual"

    if not seeds:
        print("  No seeds — nothing to do.")
        return [], ""

    print(f"\n  {len(seeds)} seeds: {seeds}")
    candidates: list[tuple[int, ...]] = []
    for s in seeds:
        periods = fam_fn(s)
        if (periods and len(periods) >= 4 and min(periods) >= 2
                and max(periods) <= 1500):
            candidates.append(periods)
    candidates = list(dict.fromkeys(candidates))
    label = f"{fam_name}×{src_name}"
    print(f"  {len(candidates)} candidate outfits generated ({label}):")
    for c in candidates[:20]:
        print(f"    {'/'.join(str(p) for p in c)}")
    if len(candidates) > 20:
        print(f"    ... and {len(candidates)-20} more")
    go = input("\nBacktest these? [Y/n]: ").strip().lower()
    if go == "n":
        return [], ""
    return candidates, label

# ── Load existing outfits to flag duplicates ──────────────────────────────────
try:
    from engine import OUTFITS
    EXISTING_PERIODS = {tuple(sorted(o["periods"])) for o in OUTFITS}
    logging.info(f"Loaded {len(EXISTING_PERIODS)} existing outfits — will flag duplicates")
except Exception:
    EXISTING_PERIODS = set()
    logging.warning("Could not import OUTFITS from engine.py — duplicate check disabled")


# ─────────────────────────────────────────────────────────────────────────────
# Phase 1 — Candidate Generation
# ─────────────────────────────────────────────────────────────────────────────

def generate_candidates(
    base_min: int   = 5,
    base_max: int   = 60,
    mult_min: float = 1.5,
    mult_max: float = 3.0,
    mult_step: float = 0.1,
    n_periods: int  = 6,
    max_period: int = 1500,
) -> list[tuple[int, ...]]:
    """
    Generate candidate 6-period SMA outfits using geometric sequences.

    For each base b and multiplier m:
        periods = [b, b*m, b*m², b*m³, b*m⁴, b*m⁵]  (rounded to integers)

    This mirrors the structure of the existing 42 outfits (e.g. 16/31/63/125/250/500
    is approximately base=16, multiplier≈2).
    """
    candidates: set[tuple[int, ...]] = set()

    def add_geometric(base: int, m: float) -> None:
        periods = tuple(sorted(set(
            int(round(base * (m ** i))) for i in range(n_periods)
        )))
        if (len(periods) == n_periods
                and min(periods) >= 2
                and max(periods) <= max_period):
            candidates.add(periods)

    # 1. Grid multipliers (as before)
    m = mult_min
    while m <= mult_max + 1e-9:
        for base in range(base_min, base_max + 1):
            add_geometric(base, m)
        m = round(m + mult_step, 10)

    # 2. Special ratios — phi, sqrt2, sqrt3, phi², sqrt(phi), perfect fifth.
    #    These fall between grid steps and dominated the Jul 9 results.
    for sm in SPECIAL_MULTS:
        for base in range(base_min, base_max + 1):
            add_geometric(base, sm)

    # 3. Fibonacci windows — 6 consecutive fib numbers × integer scale
    for start in range(len(FIB) - n_periods + 1):
        window = FIB[start:start + n_periods]
        for scale in range(1, 21):
            periods = tuple(p * scale for p in window)
            if periods[-1] <= max_period and periods[0] >= 2:
                candidates.add(periods)

    # 4. Mutations of past winners (from the novel-outfit registry).
    #    Whole-set scaling ±5/10% plus single-period nudges explores the
    #    neighbourhood of proven structures.
    for seed in load_registry_seeds():
        for scale in (0.9, 0.95, 1.05, 1.1):
            periods = tuple(sorted(set(int(round(p * scale)) for p in seed)))
            if (len(periods) == n_periods and min(periods) >= 2
                    and max(periods) <= max_period):
                candidates.add(periods)
        for i in range(len(seed)):
            for delta in (-2, -1, 1, 2):
                mutated = list(seed)
                mutated[i] = mutated[i] + delta
                periods = tuple(sorted(set(mutated)))
                if (len(periods) == n_periods and min(periods) >= 2
                        and max(periods) <= max_period):
                    candidates.add(periods)

    novel = [c for c in candidates if c not in EXISTING_PERIODS]
    logging.info(
        f"Generated {len(candidates)} candidates — "
        f"{len(novel)} novel (not in existing outfits)"
    )
    return novel


def load_registry_seeds(max_seeds: int = 40) -> list[tuple[int, ...]]:
    """Top past winners from the registry, used as mutation seeds."""
    if not REGISTRY_PATH.exists():
        return []
    seeds: list[tuple[tuple[int, ...], float]] = []
    try:
        with open(REGISTRY_PATH) as f:
            for row in csv.DictReader(f):
                try:
                    periods = tuple(int(p) for p in row["periods"].split("/"))
                    seeds.append((periods, float(row.get("best_composite", 0) or 0)))
                except (ValueError, KeyError):
                    continue
    except Exception as e:
        logging.warning(f"Registry read failed: {e}")
        return []
    seeds.sort(key=lambda x: -x[1])
    top = [s for s, _ in seeds[:max_seeds]]
    if top:
        logging.info(f"Loaded {len(top)} mutation seeds from registry")
    return top


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2 — Candle Cache
# ─────────────────────────────────────────────────────────────────────────────

def load_ticker(ticker: str, cache_dir: str = CACHE_DIR) -> dict:
    """Load candle data for one ticker. Returns {timeframe: DataFrame} or {}."""
    path = Path(cache_dir) / f"{ticker}.pkl.gz"
    if not path.exists():
        return {}
    try:
        with gzip.open(path, "rb") as f:
            return pickle.load(f)
    except Exception as e:
        logging.debug(f"  {ticker}: load error — {e}")
        return {}


# ─────────────────────────────────────────────────────────────────────────────
# Phase 3 — Touch Detection
# ─────────────────────────────────────────────────────────────────────────────

def find_touches(
    closes: np.ndarray,
    periods: tuple[int, ...],
    touch_pct: float = TOUCH_PCT,
    absence: int = ABSENCE_WIN,
) -> list[dict]:
    """
    Find SMA touch events for a candidate outfit on a close price series.

    A touch = price within touch_pct of any MA, provided price has been
    absent from that MA for at least `absence` candles beforehand.
    Same logic as the main engine's hit detection.
    """
    n = len(closes)
    max_p = max(periods)
    min_start = max_p + absence
    if n <= min_start:
        return []

    series = pd.Series(closes)
    idx = np.arange(n)
    touches = []

    for p in periods:
        sma = series.rolling(p, min_periods=p).mean().values
        with np.errstate(divide="ignore", invalid="ignore"):
            dist = np.abs(closes - sma) / sma
        near = (dist <= touch_pct) & np.isfinite(dist)

        # Absence check, vectorized: prior_any[i] = any touch of this MA in
        # the preceding `absence` candles. cumsum trick: cs[i] = touches < i.
        cs = np.concatenate(([0], np.cumsum(near.astype(np.int64))))
        lo = np.maximum(idx - absence, 0)
        prior_any = (cs[idx] - cs[lo]) > 0

        valid = near & ~prior_any & (idx >= min_start)
        for i in np.nonzero(valid)[0]:
            touches.append({"idx": int(i), "price": float(closes[i]), "period": p})

    touches.sort(key=lambda t: t["idx"])
    return touches


# ─────────────────────────────────────────────────────────────────────────────
# Phase 4 — Forward Returns
# ─────────────────────────────────────────────────────────────────────────────

def attach_forward_returns(
    closes: np.ndarray,
    touches: list[dict],
    windows: list[int] = FWD_WINDOWS,
) -> list[dict]:
    """
    Attach forward candle returns to each touch event.
    Forward return = (close[i+w] - entry) / entry × 100
    """
    n = len(closes)
    results = []
    for t in touches:
        i = t["idx"]
        entry = t["price"]
        if entry == 0:
            continue
        fwd = {w: ((closes[i + w] - entry) / entry * 100) if (i + w) < n else None
               for w in windows}
        results.append({**t, "forward": fwd})
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Phase 5 — Outfit Scoring
# ─────────────────────────────────────────────────────────────────────────────

def score_outfit(
    ticker_results: dict[str, list[dict]],
    windows: list[int] = FWD_WINDOWS,
    min_hits: int = MIN_HITS,
) -> dict | None:
    """
    Aggregate forward return performance of one outfit across all tickers.

    Only tickers with >= min_hits touch events contribute.
    Returns a score dict or None if breadth is zero.
    """
    qualifying: list[str] = []
    all_returns: dict[int, list[float]] = {w: [] for w in windows}
    oos_returns_10: list[float] = []
    top_tickers: list[tuple[str, int, float]] = []

    for ticker, events in ticker_results.items():
        if len(events) < min_hits:
            continue
        qualifying.append(ticker)
        t10 = [e["forward"][10] for e in events if e["forward"].get(10) is not None]
        top_tickers.append((ticker, len(events), sum(t10) / len(t10) if t10 else 0.0))
        for e in events:
            is_oos = e.get("oos", False)
            for w in windows:
                v = e["forward"].get(w)
                if v is None:
                    continue
                if not is_oos:
                    all_returns[w].append(v)
                elif w == 10:
                    oos_returns_10.append(v)

    breadth = len(qualifying)
    if breadth == 0:
        return None

    sc: dict = {"breadth": breadth, "tickers": ",".join(qualifying)}
    for w in windows:
        data = all_returns[w]
        if data:
            wins = sum(1 for r in data if r > 0)
            sc[f"n_{w}"]   = len(data)
            sc[f"win_{w}"] = round(wins / len(data) * 100, 1)
            sc[f"avg_{w}"] = round(sum(data) / len(data), 3)
        else:
            sc[f"n_{w}"]   = 0
            sc[f"win_{w}"] = 0.0
            sc[f"avg_{w}"] = 0.0

    # Out-of-sample stats (win10 on the held-out final segment of each series)
    if oos_returns_10:
        oos_wins = sum(1 for r in oos_returns_10 if r > 0)
        sc["oos_n_10"]   = len(oos_returns_10)
        sc["oos_win_10"] = round(oos_wins / len(oos_returns_10) * 100, 1)
        sc["oos_avg_10"] = round(sum(oos_returns_10) / len(oos_returns_10), 3)
    else:
        sc["oos_n_10"], sc["oos_win_10"], sc["oos_avg_10"] = 0, 0.0, 0.0

    sc["validated"] = ("YES" if (sc["oos_n_10"] >= OOS_MIN_N
                                 and sc["oos_win_10"] >= OOS_WIN_MIN) else "no")

    # Composite: breadth × win10_fraction × avg10 — computed on IN-SAMPLE data
    # only, so held-out performance can't leak into the ranking.
    sc["composite"] = round(
        breadth * (sc["win_10"] / 100) * max(sc["avg_10"], 0), 4
    )

    top_tickers.sort(key=lambda x: -x[1])
    sc["top_tickers"] = "|".join(f"{t}({h})" for t, h, _ in top_tickers[:5])

    return sc


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="SMA Outfit Discovery Engine")
    parser.add_argument("--tickers",    default=",".join(DEFAULT_TICKERS),
                        help="Comma-separated ticker list (default: 30-ticker basket)")
    parser.add_argument("--timeframes", default="1d,1w,1mo",
                        help="Timeframes to scan (default: 1d,1w,1mo)")
    parser.add_argument("--top",        type=int, default=25,
                        help="Top N outfits to print (default: 25)")
    parser.add_argument("--touch",      type=float, default=TOUCH_PCT,
                        help=f"Touch threshold (default: {TOUCH_PCT} = 0.3%%)")
    parser.add_argument("--absence",    type=int, default=ABSENCE_WIN,
                        help=f"Absence window in candles (default: {ABSENCE_WIN})")
    parser.add_argument("--min-hits",   type=int, default=MIN_HITS,
                        help=f"Min touches per ticker (default: {MIN_HITS})")
    parser.add_argument("--breadth",    type=float, default=BREADTH_MIN,
                        help=f"Min breadth fraction (default: {BREADTH_MIN})")
    parser.add_argument("--base-min",   type=int, default=5)
    parser.add_argument("--base-max",   type=int, default=60)
    parser.add_argument("--mult-min",   type=float, default=1.5)
    parser.add_argument("--mult-max",   type=float, default=3.0)
    parser.add_argument("--interactive", action="store_true",
                        help="Menu mode: pick a sequence family f(x) and seed source "
                             "(manual / influx cumulative-ds / entry prices / registry), "
                             "then backtest the generated outfits")
    parser.add_argument("--family", type=str, default=None,
                        choices=list(SEQUENCE_FAMILIES.keys()),
                        help="Non-interactive: sequence family for --seeds")
    parser.add_argument("--fx", type=str, default=None,
                        help="Non-interactive: custom f(x, i) expression, e.g. 'x * 1.618**i'")
    parser.add_argument("--seeds", type=str, default=None,
                        help="Non-interactive: comma-separated seed ints, or "
                             "'influx' / 'prices' / 'registry'")
    args = parser.parse_args()

    tickers    = [t.strip() for t in args.tickers.split(",") if t.strip()]
    timeframes = [tf.strip() for tf in args.timeframes.split(",") if tf.strip()]
    breadth_n  = max(1, int(len(tickers) * args.breadth))

    logging.info("=" * 60)
    logging.info("SMA OUTFIT DISCOVERY ENGINE")
    logging.info("=" * 60)
    logging.info(f"Basket ({len(tickers)}): {', '.join(tickers)}")
    logging.info(f"Timeframes: {timeframes}")
    logging.info(f"Breadth threshold: {breadth_n}/{len(tickers)} ({args.breadth*100:.0f}%)")
    logging.info(f"Touch: {args.touch*100:.2f}%  |  Absence: {args.absence} candles  |  Min hits: {args.min_hits}")

    # ── Generate candidates ───────────────────────────────────────────────────
    run_label = "grid"
    if args.interactive:
        candidates, run_label = interactive_candidates(tickers)
        if not candidates:
            return
    elif args.seeds:
        # Non-interactive seeded mode
        if args.seeds == "influx":
            seeds = seeds_from_influx_periods(tickers)
        elif args.seeds == "prices":
            seeds = seeds_from_entry_prices()
        elif args.seeds == "registry":
            seeds = seeds_from_registry()
        else:
            seeds = [int(s) for s in args.seeds.split(",") if s.strip().isdigit()]
        if args.fx:
            fam_fn, fam_name = seq_custom_fx(args.fx), f"custom({args.fx})"
        else:
            fam = args.family or "fib"
            fam_name, (_, fam_fn) = fam, SEQUENCE_FAMILIES[fam]
        candidates = []
        for s in seeds:
            periods = fam_fn(s)
            if (periods and len(periods) >= 4 and min(periods) >= 2
                    and max(periods) <= 1500):
                candidates.append(periods)
        candidates = list(dict.fromkeys(candidates))
        run_label = f"{fam_name}×{args.seeds}"
        logging.info(f"Seeded mode: {len(candidates)} candidates from "
                     f"{len(seeds)} seeds ({run_label})")
        if not candidates:
            logging.error("No valid candidates from seeds. Exiting.")
            return
    else:
        candidates = generate_candidates(
            base_min=args.base_min,
            base_max=args.base_max,
            mult_min=args.mult_min,
            mult_max=args.mult_max,
        )

    # ── Load candle cache ─────────────────────────────────────────────────────
    logging.info("Loading candle cache...")
    cache: dict[str, dict] = {}
    missing: list[str] = []
    for ticker in tickers:
        data = load_ticker(ticker)
        if data:
            cache[ticker] = data
        else:
            missing.append(ticker)

    if missing:
        logging.warning(f"Not found in cache: {missing}")
    logging.info(f"Loaded {len(cache)}/{len(tickers)} tickers")

    if not cache:
        logging.error("No tickers loaded. Exiting.")
        return

    # ── Build close arrays ────────────────────────────────────────────────────
    closes_map: dict[str, dict[str, np.ndarray]] = {}
    for ticker, tf_data in cache.items():
        closes_map[ticker] = {}
        for tf in timeframes:
            df = tf_data.get(tf)
            if df is None or df.empty or "close" not in df.columns:
                continue
            arr = df["close"].dropna().values.astype(float)
            if len(arr) > 20:
                closes_map[ticker][tf] = arr

    available_tickers = [t for t in closes_map if len(closes_map[t]) > 0]
    logging.info(f"Tickers with usable candle data: {len(available_tickers)}")

    # ── Precompute touch events per (ticker, tf, period) ─────────────────────
    # A touch of MA44 is independent of which outfit MA44 belongs to, so events
    # are computed ONCE per distinct period and candidates just assemble them.
    # This is the difference between hours and minutes on wide searches.
    from bisect import bisect_left

    distinct_periods = sorted({p for c in candidates for p in c})
    logging.info(f"Precomputing touch events: {len(distinct_periods)} distinct periods "
                 f"× {len(available_tickers)} tickers × {len(timeframes)} tfs...")

    # period_events[(ticker, tf, p)] = (sorted idx list, events list)
    period_events: dict[tuple, tuple[list[int], list[dict]]] = {}
    pre_done = 0
    for ticker in available_tickers:
        for tf in timeframes:
            arr = closes_map[ticker].get(tf)
            if arr is None:
                continue
            split_idx = int(len(arr) * OOS_FRACTION)
            for p in distinct_periods:
                pre_done += 1
                if pre_done % 20000 == 0:
                    logging.info(f"  precompute: {pre_done:,} (period, ticker, tf) done")
                if p + args.absence >= len(arr):
                    continue
                touches = find_touches(arr, (p,), args.touch, args.absence)
                if not touches:
                    continue
                evs = attach_forward_returns(arr, touches)
                if not evs:
                    continue
                for e in evs:
                    e["oos"] = e["idx"] >= split_idx
                period_events[(ticker, tf, p)] = ([e["idx"] for e in evs], evs)

    logging.info(f"Precompute done: {len(period_events):,} (ticker, tf, period) "
                 f"combos have touch events")

    # ── Assemble + score all candidates × timeframes ─────────────────────────
    logging.info(f"Scoring {len(candidates)} candidates × {len(timeframes)} timeframes = "
                 f"{len(candidates)*len(timeframes):,} combinations...")

    results: list[dict] = []
    total = len(candidates) * len(timeframes)
    done  = 0

    for candidate in candidates:
        cand_min_start = max(candidate) + args.absence
        for tf in timeframes:
            done += 1
            if done % 2000 == 0:
                pct = done / total * 100
                logging.info(f"  Progress: {done:,}/{total:,} ({pct:.1f}%) — {len(results)} passing so far")

            # Assemble this candidate's events from the per-period precompute.
            # Filter to idx >= cand_min_start — the same validity window the
            # original per-candidate scan enforced (longest MA + absence).
            ticker_results: dict[str, list[dict]] = {}
            for ticker in available_tickers:
                events: list[dict] = []
                for p in candidate:
                    pe = period_events.get((ticker, tf, p))
                    if not pe:
                        continue
                    idxs, evs = pe
                    k = bisect_left(idxs, cand_min_start)
                    if k < len(evs):
                        events.extend(evs[k:])
                if events:
                    ticker_results[ticker] = events

            # Quick breadth pre-filter before full scoring
            if len(ticker_results) < breadth_n:
                continue

            sc = score_outfit(ticker_results, min_hits=args.min_hits)
            if sc is None or sc["breadth"] < breadth_n:
                continue

            is_novel = candidate not in EXISTING_PERIODS
            results.append({
                "periods":   "/".join(str(p) for p in candidate),
                "timeframe": tf,
                "novel":     "YES" if is_novel else "NO",
                **sc,
            })

    logging.info(f"Scan complete. {len(results)} outfits met breadth threshold of {breadth_n}/{len(tickers)}")

    # ── Sort and filter ───────────────────────────────────────────────────────
    results.sort(key=lambda x: -x["composite"])

    promoted = [
        r for r in results
        if r["novel"] == "YES" and r["win_10"] >= WIN_RATE_MIN * 100
        and r.get("validated") == "YES"
    ]
    unvalidated = [
        r for r in results
        if r["novel"] == "YES" and r["win_10"] >= WIN_RATE_MIN * 100
        and r.get("validated") != "YES"
    ]

    # ── Terminal output ───────────────────────────────────────────────────────
    top_n = min(args.top, len(results))
    print(f"\n{'='*90}")
    print(f"  SMA OUTFIT DISCOVERY — Top {top_n} Results")
    print(f"  Basket: {len(cache)} tickers  |  Timeframes: {timeframes}")
    print(f"  Candidates: {len(candidates):,}  |  Passing breadth ({breadth_n}/{len(tickers)}): {len(results)}")
    print(f"  Promoted (novel + win10 ≥ {WIN_RATE_MIN*100:.0f}%): {len(promoted)}")
    print(f"{'='*90}")
    print(f"  {'Periods':<32} {'TF':<5} {'Nov':<4} {'Val':<4} {'Brd':<4} "
          f"{'W10%':<6} {'oosW10':<7} {'oosN':<5} {'Avg10':<8} "
          f"{'Score':<8} Top Tickers")
    print(f"  {'-'*32} {'-'*4} {'-'*3} {'-'*3} {'-'*3} "
          f"{'-'*5} {'-'*6} {'-'*4} {'-'*7} "
          f"{'-'*7} {'-'*25}")

    for r in results[:top_n]:
        print(
            f"  {r['periods']:<32} {r['timeframe']:<5} {r['novel']:<4} "
            f"{r.get('validated','no'):<4} "
            f"{r['breadth']:<4} {r['win_10']:<6.1f} "
            f"{r.get('oos_win_10',0):<7.1f} {r.get('oos_n_10',0):<5} "
            f"{r['avg_10']:<8.2f} "
            f"{r['composite']:<8.3f} {r.get('top_tickers','')}"
        )

    if promoted:
        print(f"\n{'='*90}")
        print(f"  ★  PROMOTED — novel, win10 ≥ {WIN_RATE_MIN*100:.0f}%, "
              f"AND validated out-of-sample (oos win10 ≥ {OOS_WIN_MIN:.0f}%, n ≥ {OOS_MIN_N})")
        print(f"{'='*90}")
        for r in promoted[:15]:
            print(
                f"  {r['periods']:<40} {r['timeframe']:<5}  "
                f"breadth={r['breadth']}  win10={r['win_10']}%  "
                f"oos={r['oos_win_10']}% (n={r['oos_n_10']})  avg10={r['avg_10']:+.2f}%"
            )
    else:
        print(f"\n  No outfits promoted at current thresholds.")
        print(f"  Try: --breadth 0.5 --absence 30 to loosen criteria.")

    if unvalidated:
        print(f"\n  ⚠  {len(unvalidated)} outfits passed in-sample but FAILED "
              f"out-of-sample validation (likely regime-fit — not promoted)")

    # ── Save CSV ──────────────────────────────────────────────────────────────
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    ts      = datetime.now(timezone.utc).strftime("%Y-%m-%d_%H-%M-%S")
    out     = OUTPUT_DIR / f"outfit_discovery_{ts}.csv"
    fields  = [
        "periods", "timeframe", "novel", "validated", "breadth", "tickers",
        "n_1",  "win_1",  "avg_1",
        "n_3",  "win_3",  "avg_3",
        "n_5",  "win_5",  "avg_5",
        "n_10", "win_10", "avg_10",
        "n_20", "win_20", "avg_20",
        "oos_n_10", "oos_win_10", "oos_avg_10",
        "composite", "top_tickers",
    ]
    with open(out, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(results)

    logging.info(f"Saved → {out}")

    # ── Update novel-outfit registry (accumulates across runs) ───────────────
    update_registry(results)

    print(f"\n  Output saved: {out}\n")


def update_registry(results: list[dict]) -> None:
    """
    Merge this run's novel outfits into all_novel_outfits.csv.
    Keyed by (periods, timeframe). Tracks first_seen, last_seen, times_seen,
    best_composite, and the latest in/out-of-sample win rates. The registry
    also feeds mutation seeds for future runs.
    """
    now = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    registry: dict[tuple[str, str], dict] = {}

    if REGISTRY_PATH.exists():
        try:
            with open(REGISTRY_PATH) as f:
                for row in csv.DictReader(f):
                    key = (row.get("periods", ""), row.get("timeframe", ""))
                    if key[0]:
                        registry[key] = row
        except Exception as e:
            logging.warning(f"Registry read failed (starting fresh): {e}")

    updated = 0
    for r in results:
        if r["novel"] != "YES":
            continue
        key = (r["periods"], r["timeframe"])
        prev = registry.get(key)
        entry = {
            "periods":         r["periods"],
            "timeframe":       r["timeframe"],
            "first_seen":      (prev.get("first_seen") or now) if prev else now,
            "last_seen":       now,
            "times_seen":      (int(prev.get("times_seen", 0)) + 1) if prev else 1,
            "best_composite":  max(float(prev.get("best_composite", 0) or 0),
                                   r["composite"]) if prev else r["composite"],
            "latest_win_10":   r["win_10"],
            "latest_oos_win_10": r.get("oos_win_10", 0.0),
            "latest_validated":  r.get("validated", "no"),
            "breadth":         r["breadth"],
        }
        registry[key] = entry
        updated += 1

    if not registry:
        return
    REGISTRY_PATH.parent.mkdir(parents=True, exist_ok=True)
    reg_fields = ["periods", "timeframe", "first_seen", "last_seen", "times_seen",
                  "best_composite", "latest_win_10", "latest_oos_win_10",
                  "latest_validated", "breadth"]
    with open(REGISTRY_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=reg_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(registry.values())
    logging.info(f"Registry updated: {updated} outfits merged, "
                 f"{len(registry)} total → {REGISTRY_PATH}")


if __name__ == "__main__":
    main()
